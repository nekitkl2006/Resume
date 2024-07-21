from pyrogram import Client, types, filters
import pyrostep
from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font
from pyrogram.types import BotCommand, BotCommandScopeAllPrivateChats
import time
from copy import copy

app = Client("Bot_account", api_id="YOUR API ID", api_hash="YOUR API HASH", bot_token="BOT TOKEN")

pyrostep.listen(app)

async def new_commands():
    bot = await app.set_bot_commands([BotCommand("basket", "Корзина🛒"), BotCommand("account", "Личный кабинет👤"), BotCommand("start", "Начать👋")], scope=BotCommandScopeAllPrivateChats())
    print(bot)

@app.on_message(filters.command(commands=['start']))
def start_bot(client, message):
    if message.chat.id == -1001789882072:
        return
    if message.from_user.id == 1333967466 or message.from_user.id == 6046959676 or message.from_user.id == 993684230:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002699Управление заказами", callback_data=f"red_table_orders")], [InlineKeyboardButton("\U0001F69BУправление закупками", callback_data=f"admin_purchage")], [InlineKeyboardButton("\U0000270FРедактировать таблицу товаров", callback_data=f"admin_table")], [InlineKeyboardButton("\U0001F440Посмотреть таблицу товаров", callback_data=f"view_table")]])
        app.send_message(chat_id=message.from_user.id, text="\U00002699Для использования функций администратора воспользуйтесь кнопками ниже\U00002B07", reply_markup=markup)
    else:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            app.send_message(chat_id=message.from_user.id, text=f"\U0001F44BBotanica_club_bot приветствует вас, **{message.from_user.first_name}**! Делайте заказы через <a href='https://t.me/botanica_club'>комментарии</a>, а дальше я помогу\U0001F917\n\n\U0001F6D2Корзину можно открыть с помощью команды /basket\n\n\U0001F464А для просмотра заказов и доступа к личному кабинету воспользуйтесь командой /account")
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0000270F\U0001F69AВвести данные для доставки", callback_data=f"add_delivery")]])
            app.send_message(chat_id=message.from_user.id, text=f"\U0001F44BBotanica_club_bot приветствует вас, **{message.from_user.first_name}**! Делайте заказы через <a href='https://t.me/botanica_club'>комментарии</a>, а дальше я помогу\U0001F917\n\n\U0001F464Воспользуйтесь кнопкой ниже, чтобы ввести данные для отправки и получения заказов", reply_markup=markup, disable_web_page_preview=True)
        connection.close()

@app.on_message(filters.command(commands=['basket']))
def basket(client, message):
    if message.chat.id == -1001789882072:
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status= ?", (message.from_user.id, 1))
    orders = cursor.fetchall()
    if orders != []:
        markup = []
        num = 0
        for order in orders:
            if num < 5:
                cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                osts = cursor.fetchall()
                button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_product{order[0]}num{order[1]}")]
                markup.append(button)
            num += 1
        if len(orders) > 5:
            markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_basket5")])
        markup.append([InlineKeyboardButton(text="\U00002705Оформить заказ\U00002705", callback_data="new_order")])
        markup.append([InlineKeyboardButton(text="📅🚚Сроки доставки", callback_data="delivery_time")])
        markup = InlineKeyboardMarkup(markup)
        app.send_message(chat_id=message.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0000203CПри заказе товаров из разных закупок доставка будет осуществляться несколькими отправлениями\n\nДля оформления заказа воспользуйтесь кнопокой\U0001F447", reply_markup=markup)
    else:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton(text="📅🚚Сроки доставки", callback_data="delivery_time")]])
        app.send_message(chat_id=message.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0001F5D1Ваша корзина пуста. Добавьте товары через комментарии!\U0001F4AC", reply_markup=markup)
    connection.close()

@app.on_message(filters.command(commands=['account']))
def my_account(client, message):
        if message.chat.id == -1001789882072:
            return
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Изменить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**.\n\n\U00002139Данные для доставки:\n**ФИО**: {user[0][1]}\n**Адрес доставки**: {user[0][2]}\n**Телефон**: {user[0][3]}"
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Заполнить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**. Воспользуйтесь кнопками ниже\U00002B07"
        app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
        connection.close()

@app.on_callback_query()
async def answer(client, function_call):
    if function_call.data == "admin_purchage":
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F195Новая закупка", callback_data=f"new_purchage")], [InlineKeyboardButton("\U00002139Посмотреть закупку", callback_data=f"info_purchage")], [InlineKeyboardButton("\U00002716Закрыть закупку", callback_data=f"del_purchage")], [InlineKeyboardButton("\U0000270FИзменить закупку", callback_data=f"edit_purchage")], [InlineKeyboardButton("\U0001F519Назад", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.message.from_user.id, text="\U0000270FЧто вы хотите сделать?", reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "new_purchage":
        await part_new_purchage(function_call)
    elif "again_new_purchage" in function_call.data:
        name = (function_call.data)[18:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("DELETE FROM purchages WHERE name = ?", (name,))
        connection.commit()
        connection.close()
        await part_new_purchage(function_call)
    elif "save_purchage" in function_call.data:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        name = (function_call.data)[13:]
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U00002705Закупка {name} успешно сохранена", message_id=function_call.message.id, reply_markup=markup)
    elif function_call.data == "exit_table":
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002699Управление заказами", callback_data=f"red_table_orders")], [InlineKeyboardButton("\U0001F69BУправление закупками", callback_data=f"admin_purchage")], [InlineKeyboardButton("\U0000270FРедактировать таблицу товаров", callback_data=f"admin_table")], [InlineKeyboardButton("\U0001F440Посмотреть таблицу товаров", callback_data=f"view_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U00002699Для использования функций администратора воспользуйтесь кнопками ниже\U00002B07", reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "del_purchage" or function_call.data == "info_purchage" or function_call.data == "edit_purchage" or function_call.data == "new_table" or function_call.data == "add_column_table" or function_call.data == "view_table" or function_call.data == "view_table_orders":
        if function_call.data == "del_purchage":
            todo = "del_name_purchage"
            next = "del_next_purchage"
            text = "\U0001F534Какую закупку нужно закрыть?\n\n\U0000203CВ случае закрытия закупки будут удалены и все заказы в рамках этой закупки. В ответ будет выслана таблица со всеми оплаченными заказами. Таблица товаров этой закупки также будет удалена"
        elif function_call.data == "info_purchage":
            todo = "inf_name_purchage"
            next = "inf_next_purchage"
            text = "\U0001F440Нажмите на закупку для просмотра информации о ней"
        elif function_call.data == "edit_purchage":
            todo = "edt_name_purchage"
            next = "edt_next_purchage"
            text = "\U0000270FНажмите на закупку для изменения информации о ней"    
        elif function_call.data == "new_table":
            todo = "new_tble_purchage"
            next = "new_next_purchage"
            text = "Нажмите на закупку, для которой нужно создать новую таблицу товаров\n\n\U0000203CВсе заказы, сделанные в рамках этой закупки, будут удалены. В ответ будет выслана таблица со всеми оплаченными заказами" 
        elif function_call.data == "add_column_table":
            todo = "cln_tble_purchage"
            next = "cln_next_purchage"
            text = "\U00002795Выберите закупку, для которой нужно добавить товар"
        elif function_call.data == "view_table":
            todo = "viw_tble_purchage"
            next = "viw_next_purchage"
            text = "\U0001F440Выберите закупку, чтобы посмотреть таблицу товаров для неё" 
        elif function_call.data == "view_table_orders":
            todo = "ord_tble_purchage"
            next = "ord_next_purchage"
            text = "\U0001F440\U0001F4E6Выберите закупку, чтобы посмотреть таблицу заказов для неё"  
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT name FROM purchages")
        purchages = cursor.fetchall()
        markup = []
        num = 0
        for purchage in purchages:
            if num < 5:
                markup.append([InlineKeyboardButton(text=f"{purchage[0]}", callback_data=f"{todo}{purchage[0]}")])
            num += 1
        if len(purchages) > 5:
            markup.append([InlineKeyboardButton(text="\U000027A1", callback_data=f"{next}5")])
        if function_call.data == "del_purchage" or function_call.data == "info_purchage" or function_call.data == "edit_purchage":
            markup.append([InlineKeyboardButton(text="\U0001F519Назад", callback_data=f"admin_purchage")])
        elif function_call.data == "new_table" or function_call.data == "add_column_table":
            markup.append([InlineKeyboardButton(text="\U0001F519Назад", callback_data=f"admin_table")])
        elif function_call.data == "view_table":
            markup.append([InlineKeyboardButton(text="\U0001F519Назад", callback_data=f"exit_table")])
        markup = InlineKeyboardMarkup(markup)
        connection.close()
        await app.edit_message_text(chat_id=function_call.from_user.id, text=text, message_id=function_call.message.id, reply_markup=markup)
    elif "del_next_purchage" in function_call.data or "edt_next_purchage" in function_call.data or "inf_next_purchage" in function_call.data or "new_next_purchage" in function_call.data or "cln_next_purchage" in function_call.data or "viw_next_purchage" in function_call.data or "ord_next_purchage" in function_call.data:
        next = (function_call.data)[17:]
        next = int(next)
        if "del_next_purchage" in function_call.data:
            todo = "del_name_purchage"
            next_name = "del_next_purchage"
            text = "\U0001F534Какую закупку нужно закрыть?\n\n\U0000203CВ случае закрытия закупки будут удалены и все заказы в рамках этой закупки. В ответ будет выслана таблица со всеми оплаченными заказами. Таблица товаров этой закупки также будет удалена"
        elif "inf_next_purchage" in function_call.data:
            todo = "inf_name_purchage"
            next_name = "inf_next_purchage"
            text = "\U0001F440Нажмите на закупку для просмотра информации о ней"
        elif "edt_next_purchage" in function_call.data:
            todo = "edt_name_purchage"
            next_name = "edt_next_purchage"
            text = "\U0000270FНажмите на закупку для изменения информации о ней"  
        elif "new_next_purchage" in function_call.data:
            todo = "new_tble_purchage"
            next_name = "new_next_purchage"
            text = "Нажмите на закупку, для которой нужно создать новую таблицу товаров\n\n\U0000203CВсе заказы, сделанные в рамках этой закупки, будут удалены. В ответ будет выслана таблица со всеми оплаченными заказами"   
        elif "cln_next_purchage" in function_call.data:
            todo = "cln_tble_purchage"
            next_name = "cln_next_purchage"
            text = "\U00002795Выберите закупку, для которой нужно добавить товар" 
        elif "viw_next_purchage" in function_call.data:
            todo = "viw_tble_purchage"
            next_name = "viw_next_purchage"
            text = "\U0001F440Выберите закупку, чтобы посмотреть таблицу товаров для неё"
        elif "ord_next_purchage" in function_call.data:
            todo = "ord_tble_purchage"
            next_name = "ord_next_purchage"
            text = "\U0001F440\U0001F4E6Выберите закупку, чтобы посмотреть таблицу заказов для неё"  
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT name FROM purchages")
        purchages = cursor.fetchall()
        markup = []
        num = 0
        for purchage in purchages:
            if num >= next and num < next + 5:
                markup.append([InlineKeyboardButton(text=f"{purchage[0]}", callback_data=f"{todo}{purchage[0]}")])
            num += 1
        button = []
        if next >= 5:
            button.append(InlineKeyboardButton(text="\U00002B05", callback_data=f"{next_name}{next - 5}"))
        if len(purchages) - next > 5:
            button.append(InlineKeyboardButton(text="\U000027A1", callback_data=f"{next_name}{next + 5}"))
        markup.append(button)
        if "del_next_purchage" in function_call.data or "edt_next_purchage" in function_call.data or "inf_next_purchage" in function_call.data:
            markup.append([InlineKeyboardButton(text="\U0001F519Назад", callback_data=f"admin_purchage")])
        elif "new_next_purchage" in function_call.data or "cln_next_purchage" in function_call.data:
            markup.append([InlineKeyboardButton(text="\U0001F519Назад", callback_data=f"admin_table")])
        elif "viw_next_purchage" in function_call.data:
            markup.append([InlineKeyboardButton(text="\U0001F519Назад", callback_data=f"exit_table")])
        markup = InlineKeyboardMarkup(markup)
        connection.close()
        await app.edit_message_text(chat_id=function_call.from_user.id, text=text, message_id=function_call.message.id, reply_markup=markup)
    elif "del_name_purchage" in function_call.data:
        name = (function_call.data)[17:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        connection.commit()
        connection.close()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        connection.close()
        await view_table_orders(function_call, name)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("DELETE FROM purchages WHERE name = ?", (name,))
        connection.commit()
        cursor.execute("DELETE FROM Plants WHERE purchage = ?", (name,))
        connection.commit()
        cursor.execute("DELETE FROM orders WHERE purchage = ?", (name,))
        await app.send_message(chat_id=function_call.from_user.id, text=f"\U00002705Закупка {name} успешно закрыта", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
    elif "inf_name_purchage" in function_call.data:
        name = (function_call.data)[17:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT name, time1, time2, time3, time4, time5 FROM purchages WHERE name = ?", (name,))
        purchage = cursor.fetchall()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U00002139Информация о закупке:\n\n**Название:** {purchage[0][0]}\n**Обработка заказа:** {purchage[0][1]}\n**Отправка заказа садовнику:** {purchage[0][2]}\n**Оплата заказа:** {purchage[0][3]}\n**Поступление растений:** {purchage[0][4]}\n**Отправка растений:** {purchage[0][5]}", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "edt_name_purchage" in function_call.data:
        name = (function_call.data)[17:]
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FВведите сроки закупки в формате:\n\nОбработка заказа; Отправка заказа садовнику; Оплата заказа; Поступление растений; Отправка растений\n\n\U00002705Пример:\n1-2 дней; 3-5 дней; 7 дней; 5-7 дней; 8-10 часов", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, edit_purchage, kwargs={"name": name})
    elif function_call.data == "admin_table":
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F195Новая таблица", callback_data=f"new_table")], [InlineKeyboardButton("\U00002796Удалить товар", callback_data=f"del_column_table")], [InlineKeyboardButton("\U00002795Добавить товар", callback_data=f"add_column_table")], [InlineKeyboardButton("\U0001F4CFИзменить количество", callback_data=f"count_table")], [InlineKeyboardButton("\U0001F519Назад", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FЧто вы хотите сделать?", reply_markup=markup, message_id=function_call.message.id)
    elif "new_tble_purchage" in function_call.data:
        name = (function_call.data)[17:]
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        await app.send_message(chat_id=function_call.from_user.id, text=f"\U0001F195Введите таблицу для закупки **{name}** одним сообщением в формате **Номер товара - Название товара - Количество товара - Цена товара**\n\n\U00002705Пример: 25 - Антуриум - 38 - 700, 35 - Сизигиум - 5 - 800, ...", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, new_table, kwargs={"name": name})
    elif "right_new_table" in function_call.data:
        name = (function_call.data)[15:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, purchage, name, count, price FROM Temp_plants")
        all_plants = cursor.fetchall()
        cursor.execute("DELETE FROM Plants WHERE purchage = ?", (name,))
        connection.commit()
        for plant in all_plants:
            cursor.execute("INSERT INTO Plants (num, purchage, name, count, price, book, ordered) VALUES (?, ?, ?, ?, ?, ?, ?)", (plant[0], plant[1], plant[2], plant[3], plant[4], 0, 0))
        connection.commit()
        connection.close()
        await view_table_orders(function_call, name)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("DELETE FROM orders WHERE purchage = ?", (name,))
        connection.commit()
        connection.close()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Таблица для закупки {all_plants[0][1]} сохранена\U00002705", message_id=function_call.message.id, reply_markup=markup)
    elif "cln_tble_purchage" in function_call.data:
        name = (function_call.data)[17:]
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        await app.send_message(chat_id=function_call.from_user.id, text="\U00002795Введите информацию о новом товаре в формате **Номер товара - Название товара - Количество товара - Цена товара**\n\n\U00002705Пример: 25 - Антуриум - 38 - 700", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, add_column_table, kwargs={"name": name})
    elif function_call.data == "del_column_table":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        await app.send_message(chat_id=function_call.from_user.id, text="\U0001F522Введите номер товара, который нужно **удалить**", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, del_column_table)
    elif function_call.data == "count_table":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        await app.send_message(chat_id=function_call.from_user.id, text="\U0001F4CFВведите номер товара, количество которого нужно изменить", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, count_table)
    elif "viw_tble_purchage" in function_call.data:
        name = (function_call.data)[17:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, name, count, price, book, ordered FROM Plants WHERE purchage = ?", (name,))
        all_plants = cursor.fetchall()
        text = f"\U0001F440Таблица для закупки **{name}**:\n\n**Номер | Название растения | Количество | Цена | Забронировано | Заказано**\n"
        for plant in all_plants:
            text += f"{plant[0]} | {plant[1]} | {plant[2]} | {plant[3]}р. | {plant[4]} | {plant[5]}\n"
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=text, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "info_product" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[12:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        button = []
        if order != []:
            if order[0][2] > 1:
                button.append(InlineKeyboardButton(text="\U00002796", callback_data=f"minus_product{num_order}num{num_product}"))
            cursor.execute("SELECT count, price, purchage FROM Plants WHERE num = ?", (int(num_product),))
            count = cursor.fetchall()
            cursor.execute("SELECT time1, time2, time3, time4, time5 FROM purchages WHERE name = ?", (count[0][2],))
            times = cursor.fetchall()
            if count[0][0] > 0:
                button.append(InlineKeyboardButton(text="\U00002795", callback_data=f"plus_product{num_order}num{num_product}"))
            markup = [button]
            markup.append([InlineKeyboardButton(text="\U0000274CУдалить товар", callback_data=f"delete_product{num_order}num{num_product}")])
            markup.append([InlineKeyboardButton(text="\U0001F519\U0001F6D2Корзина", callback_data="basket")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"**{order[0][1]}**, {order[0][2]}шт. {order[0][2] * count[0][1]}₽\n\n📅🚚Сроки доставки:\n\U0001F4CDОбработка заказа: {times[0][0]}\n\U0001F4CDОтправка заказа садовнику: {times[0][1]}\n\U0001F4CDОплата заказа: {times[0][2]}\n\U0001F4CDПоступление растений: {times[0][3]}\n\U0001F4CDОтправка растений: {times[0][4]}", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton(text="\U0001F519\U0001F6D2Корзина", callback_data="basket")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0000274CНе удалось открыть товар. Попробуйте снова", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "next_basket" in function_call.data:
        next_b = (function_call.data)[11:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 1))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num >= int(next_b) and num < int(next_b) + 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_product{order[0]}num{order[1]}")]
                    markup.append(button)
                num += 1
            if int(next_b) >= 5:
                markup.append([InlineKeyboardButton(text="Назад\U000025C0", callback_data=f"next_basket{int(next_b) - 5}")])
            if len(orders) - int(next_b) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_basket{int(next_b) + 5}")])
            markup.append([InlineKeyboardButton(text="\U00002705Оформить заказ\U00002705", callback_data="new_order")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0000203CПри заказе товаров из разных закупок доставка будет осуществляться несколькими отправлениями\n\nДля оформления заказа воспользуйтесь кнопокой\U0001F447", reply_markup=markup, message_id=function_call.message.id)
        else:
            await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
            await app.send_message(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0001F5D1Ваша корзина пуста. Добавьте товары через комментарии!\U0001F4AC")
        connection.close()
    elif function_call.data == "basket":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 1))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num < 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_product{order[0]}num{order[1]}")]
                    num += 1
                    markup.append(button)
            if len(orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_basket5")])
            markup.append([InlineKeyboardButton(text="\U00002705Оформить заказ\U00002705", callback_data="new_order")])
            markup.append([InlineKeyboardButton(text="📅🚚Сроки доставки", callback_data="delivery_time")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0000203CПри заказе товаров из разных закупок доставка будет осуществляться несколькими отправлениями\n\nДля оформления заказа воспользуйтесь кнопокой\U0001F447", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton(text="📅🚚Сроки доставки", callback_data="delivery_time")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0001F5D1Ваша корзина пуста. Добавьте товары через комментарии!\U0001F4AC", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif  "plus_product" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[12:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        cursor.execute("SELECT time1, time2, time3, time4, time5 FROM purchages WHERE name = ?", (order[0][6],))
        times = cursor.fetchall()
        cursor.execute("UPDATE orders SET count = ? WHERE num = ? and product = ?", (order[0][2] + 1, int(num_order), int(num_product)))
        connection.commit()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        cursor.execute("SELECT book FROM Plants WHERE num = ?", (int(num_product),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] + 1, int(num_product)))
        connection.commit()
        button = []
        if order[0][2] > 1:
            button.append(InlineKeyboardButton(text="\U00002796", callback_data=f"minus_product{num_order}num{num_product}"))
        cursor.execute("SELECT count, price, book FROM Plants WHERE num = ?", (int(num_product),))
        count = cursor.fetchall()
        if count[0][0] > 0:
            button.append(InlineKeyboardButton(text="\U00002795", callback_data=f"plus_product{num_order}num{num_product}"))
        markup = [button]
        markup.append([InlineKeyboardButton(text="\U0000274CУдалить товар", callback_data=f"delete_product{num_order}num{num_product}")])
        markup.append([InlineKeyboardButton(text="\U0001F519\U0001F6D2Корзина", callback_data="basket")])
        markup = InlineKeyboardMarkup(markup)
        connection.close()
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"**{order[0][1]}**, {order[0][2]}шт. {order[0][2] * count[0][1]}₽\n\n📅🚚Сроки доставки:\n\U0001F4CDОбработка заказа: {times[0][0]}\n\U0001F4CDОтправка заказа садовнику: {times[0][1]}\n\U0001F4CDОплата заказа: {times[0][2]}\n\U0001F4CDПоступление растений: {times[0][3]}\n\U0001F4CDОтправка растений: {times[0][4]}", reply_markup=markup, message_id=function_call.message.id)
    elif  "minus_product" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[13:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        cursor.execute("SELECT time1, time2, time3, time4, time5 FROM purchages WHERE name = ?", (order[0][6],))
        times = cursor.fetchall()
        cursor.execute("UPDATE orders SET count = ? WHERE num = ? and product = ?", (order[0][2] - 1, int(num_order), int(num_product)))
        connection.commit()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        button = []
        if order[0][2] > 1:
            button.append(InlineKeyboardButton(text="\U00002796", callback_data=f"minus_product{num_order}num{num_product}"))
        cursor.execute("SELECT book FROM Plants WHERE num = ?", (int(num_product),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - 1, int(num_product)))
        connection.commit()
        cursor.execute("SELECT count, price, book FROM Plants WHERE num = ?", (int(num_product),))
        count = cursor.fetchall()
        if count[0][0] > 0:
            button.append(InlineKeyboardButton(text="\U00002795", callback_data=f"plus_product{num_order}num{num_product}"))
        markup = [button]
        markup.append([InlineKeyboardButton(text="\U0000274CУдалить товар", callback_data=f"delete_product{num_order}num{num_product}")])
        markup.append([InlineKeyboardButton(text="\U0001F519\U0001F6D2Корзина", callback_data="basket")])
        markup = InlineKeyboardMarkup(markup)
        connection.close()
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"**{order[0][1]}**, {order[0][2]}шт. {order[0][2] * count[0][1]}₽\n\n📅🚚Сроки доставки:\n\U0001F4CDОбработка заказа: {times[0][0]}\n\U0001F4CDОтправка заказа садовнику: {times[0][1]}\n\U0001F4CDОплата заказа: {times[0][2]}\n\U0001F4CDПоступление растений: {times[0][3]}\n\U0001F4CDОтправка растений: {times[0][4]}", reply_markup=markup, message_id=function_call.message.id)
    elif "delete_product" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[14:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT count FROM orders WHERE num = ? and product = ?", (int(num_order), int(num_product)))
        count = cursor.fetchall()
        cursor.execute("SELECT book FROM Plants WHERE num = ?", (int(num_product),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - count[0][0], int(num_product)))
        connection.commit()
        cursor.execute("DELETE FROM orders WHERE num = ? and product = ?", (int(num_order), int(num_product)))
        connection.commit()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 1))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num < 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_product{order[0]}num{order[1]}")]
                    markup.append(button)
                num += 1
            if len(orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_basket5")])
            markup.append([InlineKeyboardButton(text="\U00002705Оформить заказ\U00002705", callback_data="new_order")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0000203CПри заказе товаров из разных закупок доставка будет осуществляться несколькими отправлениями\n\nДля оформления заказа воспользуйтесь кнопокой\U0001F447", reply_markup=markup, message_id=function_call.message.id)
        else:
            await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
            await app.send_message(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0001F5D1Ваша корзина пуста. Добавьте товары через комментарии!\U0001F4AC") 
        connection.close()
    elif function_call.data == "red_table_orders":
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F195Новый заказ", callback_data=f"add_admin_order")], [InlineKeyboardButton("\U0001F440Посмотреть заказ", callback_data=f"view_order")], [InlineKeyboardButton("\U0000274CУдалить заказ", callback_data=f"del_order")], [InlineKeyboardButton("\U0001F440\U0001F4CEПосмотреть таблицу заказов", callback_data=f"view_table_orders")], [InlineKeyboardButton("\U0000274C\U0001F4B8Удалить неоплаченные заказы", callback_data=f"delete_unpay")], [InlineKeyboardButton("\U0001F519Назад", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FЧто вы хотите сделать?", reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "new_order":
        try:
            connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
            cursor = connection.cursor()
            cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 1))
            orders = cursor.fetchall()
            text = f"\U0001F7E2\U0001F195Новый заказ №`{orders[0][0]}` от пользователя {orders[0][5]}!\n\n"
            order_num = orders[0][0]
            user_text = f"\U0001F7E2\U0001F551Ожидайте подтверждение заказа №`{orders[0][0]}` продавцом\n\n\U0001F4E6Состав заказа и сроки доставки:\n\n"
            spis_purchages = []
            for order in orders:
                if order[8] not in spis_purchages:
                    spis_purchages.append(order[8])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                user_text += f"**Закупка {purchage}**\n"
                text += f"**Закупка {purchage}**\n"
                num = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (order_num, purchage))
                orders = cursor.fetchall()
                for order in orders:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (order[1],))
                    price = cursor.fetchall()
                    text += f"{num}. Товар {order[1]} **{order[2]}**. Кол-во: {order[3]} на сумму {price[0][0] * order[3]}₽.\n"
                    user_text += f"{num}. Товар {order[1]} **{order[2]}**. Кол-во: {order[3]} на сумму {price[0][0] * order[3]}₽.\n"
                    cost += price[0][0] * order[3]
                    num += 1
            cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (2, orders[0][0]))
            connection.commit()
            text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            user_text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            markup = InlineKeyboardMarkup([[InlineKeyboardButton(text="\U00002705Подтвердить заказ", callback_data=f"accept_all_order{orders[0][0]}")], [InlineKeyboardButton(text="\U00002611Подтвердить часть", callback_data=f"accept_part_order{orders[0][0]}")], [InlineKeyboardButton(text="\U0000274CОтклонить заказ", callback_data=f"discard_order{orders[0][0]}")]])
            connection.close()
            await app.send_message(chat_id=1333967466, text=text, reply_markup=markup)
            await app.send_message(chat_id=993684230, text=text, reply_markup=markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text=user_text + "\n\n📅🚚Для просмотра сроков доставки используйте команду /basket, а далее соответствующую кнопку", message_id=function_call.message.id)
        except:
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0000274Cпроизошла ошибка. Попробуйте снова с помощью команды /basket", message_id=function_call.message.id)
    elif "discard_order" in function_call.data:
        num = (function_call.data)[13:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        if orders != []:
            text = f"\U0001F534\U0001F6D2Продавец полностью отклонил возможность осуществления заказа №`{num}`\n\n\U0001F4E6Состав заказа:\n"
            spis_purchages = []
            for order in orders:
                if order[5] not in spis_purchages:
                    spis_purchages.append(order[5])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                numm = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                orderss = cursor.fetchall()
                for orderr in orderss:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (orderr[1],))
                    price = cursor.fetchall()
                    text += f"{numm}. Товар {orderr[1]} **{orderr[2]}**. Кол-во: {orderr[3]} на сумму {price[0][0] * orderr[3]}₽.\n"
                    cost += price[0][0] * orderr[3]
                    numm += 1
            text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            for order in orders:
                cursor.execute("SELECT book FROM Plants WHERE num = ?", (order[0],))
                book = cursor.fetchall()
                cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - order[2], order[0]))
            cursor.execute("DELETE FROM orders WHERE num = ?", (num,))
            connection.commit()
            await app.send_message(chat_id=orders[0][3], text=text)
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0001F534\U0001F6D2Заказ №`{num}` успешно отклонён", message_id=function_call.message.id)
        else:
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Пользователь уже отменил заказ или забронированный товар был удалён из списка товаров", message_id=function_call.message.id)
        connection.close()
    elif "accept_part_order" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE num = ?", ((function_call.data)[17:],))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num < 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_redactt{order[0]}num{order[1]}")]
                    num += 1
                    markup.append(button)
            if len(orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_redact5num{(function_call.data)[17:]}")])
            markup.append([InlineKeyboardButton(text="\U00002705Подтвердить заказ", callback_data=f"end_accept_part{(function_call.data)[17:]}")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U00002699Воспользуйтесь кнопками для изменения количества товаров\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Пользователь отменил заказ или забронированный товар был удалён из списка товаров", message_id=function_call.message.id)
        connection.close()
    elif "info_redactt" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[12:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        button = []
        if order[0][2] > 1:
            button.append(InlineKeyboardButton(text="\U00002796", callback_data=f"minus_redactt{num_order}num{num_product}"))
        cursor.execute("SELECT count, price, book FROM Plants WHERE num = ?", (int(num_product),))
        count = cursor.fetchall()
        if (count[0][0] - count[0][2]) > 0:
            button.append(InlineKeyboardButton(text="\U00002795", callback_data=f"plus_redactt{num_order}num{num_product}"))
        markup = [button]
        markup.append([InlineKeyboardButton(text="\U0000274CУдалить товар", callback_data=f"delete_redactt{num_order}num{num_product}")])
        markup.append([InlineKeyboardButton(text="\U0001F519Список товаров", callback_data=f"accept_part_order{num_order}")])
        markup = InlineKeyboardMarkup(markup)
        connection.close()
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"**{order[0][1]}, {order[0][2]}шт. {order[0][2] * count[0][1]}₽\n\nЗакупка: **{order[0][6]}**", reply_markup=markup, message_id=function_call.message.id)
    elif "plus_redactt" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[12:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        cursor.execute("UPDATE orders SET count = ? WHERE num = ? and product = ?", (order[0][2] + 1, int(num_order), int(num_product)))
        connection.commit()
        cursor.execute("SELECT book FROM Plants WHERE num = ?", (int(num_product),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] + 1, int(num_product)))
        connection.commit()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        button = []
        if order[0][2] > 1:
            button.append(InlineKeyboardButton(text="\U00002796", callback_data=f"minus_redactt{num_order}num{num_product}"))
        cursor.execute("SELECT count, price, book FROM Plants WHERE num = ?", (int(num_product),))
        count = cursor.fetchall()
        if (count[0][0] - count[0][2]) > 0:
            button.append(InlineKeyboardButton(text="\U00002795", callback_data=f"plus_redactt{num_order}num{num_product}"))
        markup = [button]
        connection.close()
        markup.append([InlineKeyboardButton(text="\U0000274CУдалить товар", callback_data=f"delete_redactt{num_order}num{num_product}")])
        markup.append([InlineKeyboardButton(text="\U0001F519Список товаров", callback_data=f"accept_part_order{num_order}")])
        markup = InlineKeyboardMarkup(markup)
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"**{order[0][1]}, {order[0][2]}шт. {order[0][2] * count[0][1]}₽\n\nЗакупка: **{order[0][6]}**", reply_markup=markup, message_id=function_call.message.id)
    elif "minus_redactt" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[13:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        cursor.execute("UPDATE orders SET count = ? WHERE num = ? and product = ?", (order[0][2] - 1, int(num_order), int(num_product)))
        connection.commit()
        cursor.execute("SELECT book FROM Plants WHERE num = ?", (int(num_product),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - 1, int(num_product)))
        connection.commit()
        cursor.execute('SELECT num, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ? and product = ?', (int(num_order), int(num_product)))
        order = cursor.fetchall()
        button = []
        if order[0][2] > 1:
            button.append(InlineKeyboardButton(text="\U00002796", callback_data=f"minus_redactt{num_order}num{num_product}"))
        cursor.execute("SELECT count, price, book FROM Plants WHERE num = ?", (int(num_product),))
        count = cursor.fetchall()
        if (count[0][0] - count[0][2]) > 0:
            button.append(InlineKeyboardButton(text="\U00002795", callback_data=f"plus_redactt{num_order}num{num_product}"))
        markup = [button]
        connection.close()
        markup.append([InlineKeyboardButton(text="\U0000274CУдалить товар", callback_data=f"delete_redactt{num_order}num{num_product}")])
        markup.append([InlineKeyboardButton(text="\U0001F519Список товаров", callback_data=f"accept_part_order{num_order}")])
        markup = InlineKeyboardMarkup(markup)
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"**{order[0][1]}, {order[0][2]}шт. {order[0][2] * count[0][1]}₽\n\nЗакупка: **{order[0][6]}**", reply_markup=markup, message_id=function_call.message.id)
    elif "delete_redactt" in function_call.data:
        ind = (function_call.data).find("num")
        num_order = (function_call.data)[14:ind]
        num_product = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT count FROM orders WHERE num = ? and product = ?", (int(num_order), int(num_product)))
        count = cursor.fetchall()
        cursor.execute("SELECT book FROM Plants WHERE num = ?", (int(num_product),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - count[0][0], int(num_product)))
        connection.commit()
        cursor.execute("DELETE FROM orders WHERE num = ? and product = ?", (int(num_order), int(num_product)))
        connection.commit()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ?", (int(num_order),))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num < 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_redactt{order[0]}num{order[1]}")]
                    num += 1
                    markup.append(button)
            if len(orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_redact5num{num_order}")])
            markup.append([InlineKeyboardButton(text="\U00002705Подтвердить заказ", callback_data=f"end_accept_part{num_order}")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U00002699Воспользуйтесь кнопками для изменения количества товаров\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "next_redact" in function_call.data:
        ind = (function_call.data).find("num")
        next_b = (function_call.data)[11:ind]
        num_order = (function_call.data)[(ind + 3):]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status, purchage FROM orders WHERE num = ?", (num_order,))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num >= int(next_b) and num < int(next_b) + 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽\n\nЗакупка: **{order[0][6]}**", callback_data=f"info_redactt{order[0]}num{order[1]}")]
                    markup.append(button)
                num += 1
            if int(next_b) >= 5:
                markup.append([InlineKeyboardButton(text="Назад\U000025C0", callback_data=f"next_redact{int(next_b) - 5}num{num_order}")])
            if len(orders) - int(next_b) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_redact{int(next_b) + 5}num{num_order}")])
            markup.append([InlineKeyboardButton(text="\U00002705Подтвердить заказ", callback_data=f"end_accept_part{(function_call.data)[14:]}")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U00002699Воспользуйтесь кнопками для изменения количества товаров\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "add_delivery":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000203CЗапрашиваемая информация нужна исключительно для отправки ваших заказов\n\n\U0000270FВведите ваше ФИО", reply_markup=markup, message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, add_lk_name)
    elif function_call.data == "my_orders":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        text = "\U0001F4E6Список ваших заказов:\n\n"
        cursor.execute("SELECT num, product_name, count, is_pay, status, product FROM orders WHERE customer_id = ?", (function_call.from_user.id,))
        orders = cursor.fetchall()
        spis_orders = []
        for order in orders:
            if order[0] not in spis_orders:
                spis_orders.append(order[0])
        if orders != []:
            for number in spis_orders:
                cursor.execute("SELECT num, product_name, count, is_pay, status, product, purchage FROM orders WHERE num = ?", (number,))
                ord = cursor.fetchall()
                text += f"\U0001F4CD**Заказ №{ord[0][0]}**:\n"
                ord_num = ord[0][0]
                if ord[0][3] == 1:
                    pay = "Оплачен"
                else:
                    pay = "Не оплачен"
                cursor.execute("SELECT price FROM Plants WHERE num = ?", (ord[0][5],))
                price = cursor.fetchall()
                if ord[0][4] == 1:
                    status = "В корзине\U0001F6D2"
                elif ord[0][4] == 2:
                    status = "Ожидает проверки наличия товара\U0001F50E"
                elif ord[0][4] == 3:
                    status = "Ожидает оплаты\U0001F4B3"
                elif ord[0][4] == 4:
                    status = "Ожидает проверки оплаты\U0001F551\U0001F4B5"
                elif ord[0][4] == 5:
                    status = "Ожидает ввода данных для отправки\U0001F4E8"
                elif ord[0][4] == 6:
                    status = "Оплата подтверждена, заказ принят\U00002705"
                text += f"**Оплата:** {pay}\n**Статус:** {status}\n"
                spis_purchages = []
                for one_ord in ord:
                    if one_ord[6] not in spis_purchages:
                        spis_purchages.append(one_ord[6])
                cost = 0
                spis_purchages = sorted(spis_purchages)
                for purchage in spis_purchages:
                    text += f"**Закупка {purchage}**\n"
                    num = 1
                    cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (ord_num, purchage))
                    ords = cursor.fetchall()
                    for one_ord in ords:
                        cursor.execute("SELECT price FROM Plants WHERE num = ?", (one_ord[1],))
                        price = cursor.fetchall()
                        text += f"{num}. Товар {one_ord[1]} **{one_ord[2]}**. Кол-во: {one_ord[3]} на сумму {price[0][0] * one_ord[3]}₽.\n"
                        cost += price[0][0] * one_ord[3]
                        num += 1
                text += f"\U0001F4B5Общая стоимость: {cost}₽\n\n"
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F4B5Оплатить заказ", callback_data=f"pay_my_order")], [InlineKeyboardButton("\U0000274CОтменить заказ", callback_data=f"del_my_order")], [InlineKeyboardButton("\U0001F5F3\U00002699Изменить способ получения заказа", callback_data=f"change_recieve_order")], [InlineKeyboardButton(text="📅🚚Сроки доставки", callback_data="delivery_time_account")], [InlineKeyboardButton("\U0001F519Назад", callback_data=f"back_to_account")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=text, reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"back_to_account")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F4E6\U0001F5D1Список ваших заказов пуст", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "back_to_account":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (function_call.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Изменить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**.\n\n\U00002139Данные для доставки:\n**ФИО**: {user[0][1]}\n**Адрес доставки**: {user[0][2]}\n**Телефон**: {user[0][3]}"
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Заполнить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**. Воспользуйтесь кнопками ниже\U00002B07"
        await app.edit_message_text(chat_id=function_call.from_user.id, text=text, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "change_my_data":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F3E0Главное меню")]], resize_keyboard=True)
        await app.send_message(chat_id=function_call.from_user.id, text="\U0000203CЗапрашиваемая информация нужна исключительно для отправки ваших заказов\n\n\U0000270FВведите ваше ФИО", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, add_lk_name)
    elif function_call.data == "del_my_order":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and is_pay IS NULL", (function_call.from_user.id,))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            spis_orders = []
            for order in orders:
                if order[0] not in spis_orders:
                    spis_orders.append(order[0])
            spis_orders = sorted(spis_orders)
            for ord in spis_orders:
                if num < 5:
                    button = [InlineKeyboardButton(text=f"Заказ №{ord}", callback_data=f"delete_order{ord}")]
                    num += 1
                    markup.append(button)
            if len(spis_orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"delete_next_order5")])
            markup.append([InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CВыберите заказ, который нужно отменить\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОтменить можно только неоплаченный заказ.\n\nНеоплаченные заказы отсутствуют", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "delete_next_order" in function_call.data:
        next_b = (function_call.data)[17:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and is_pay IS NULL", (function_call.from_user.id,))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            spis_orders = []
            for order in orders:
                if order[0] not in spis_orders:
                    spis_orders.append(order[0])
            spis_orders = sorted(spis_orders)
            for ord in spis_orders:
                if num >= int(next_b) and num < int(next_b) + 5:
                    button = [InlineKeyboardButton(text=f"Заказ №{ord}", callback_data=f"delete_order{ord}")]
                    markup.append(button)
                num += 1
            if int(next_b) >= 5:
                markup.append([InlineKeyboardButton(text="Назад\U000025C0", callback_data=f"delete_next_order{int(next_b) - 5}")])
            if len(spis_orders) - int(next_b) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"delete_next_order{int(next_b) + 5}")])
            markup.append([InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CВыберите заказ, который нужно отменить\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОтменить можно только неоплаченный заказ.\n\nНеоплаченные заказы отсутствуют", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif "delete_order" in function_call.data:
        num = (function_call.data)[12:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, reciept, is_pay FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        for order in orders:
            cursor.execute("SELECT book FROM Plants WHERE num = ?", (order[1],))
            count = cursor.fetchall()
            cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (count[0][0] - order[3], order[1]))
            connection.commit()
        cursor.execute("DELETE FROM orders WHERE num = ?", (num,))
        connection.commit()
        connection.close()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"back_to_account")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U00002705Заказ №`{orders[0][0]}` успешно отменён", reply_markup=markup, message_id=function_call.message.id)
    elif "accept_all_order" in function_call.data:
        num = (function_call.data)[16:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        if orders != []:
            cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (3, num))
            connection.commit()
            text = f"\U00002714Продавец подтвердил заказ №`{num}`\n\n\U0001F4E6Состав заказа:\n"
            spis_purchages = []
            for order in orders:
                if order[5] not in spis_purchages:
                    spis_purchages.append(order[5])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                number = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                orders = cursor.fetchall()
                for order in orders:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (order[1],))
                    price = cursor.fetchall()
                    text += f"{number}. Товар {order[1]} **{order[2]}**. Кол-во: {order[3]} на сумму {price[0][0] * order[3]}₽.\n"
                    cost += price[0][0] * order[3]
                    number += 1
            text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            cursor.execute("SELECT card_data FROM Payment")
            card_data = cursor.fetchall()
            text += f"\n\n\U0001F4B3Оплатите заказ по реквизитам:\n{card_data[0][0]}"
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Я оплатил", callback_data=f"buy_done{num}")], [InlineKeyboardButton("\U0000274CОтказаться от оплаты", callback_data=f"buy_user_reject{num}")]])
            await app.send_message(chat_id=orders[0][4], text=text, reply_markup=markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0001F4E8Заказ №`{num}` подтверждён\U00002705", message_id=function_call.message.id)
        else:
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Пользователь отменил заказ или забронированный товар был удалён из списка товаров", message_id=function_call.message.id)
        connection.close()
    elif "end_accept_part" in function_call.data:
        num = (function_call.data)[15:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        if orders != []:
            cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (3, num))
            connection.commit()
            text = f"\U00002714\U0000203CПродавец подтвердил только часть заказа №`{num}`\n\n\U0001F4E6Новый состав заказа:\n"
            spis_purchages = []
            for order in orders:
                if order[5] not in spis_purchages:
                    spis_purchages.append(order[5])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                number = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                orders = cursor.fetchall()
                for order in orders:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (order[1],))
                    price = cursor.fetchall()
                    text += f"{number}. Товар {order[1]} **{order[2]}**. Кол-во: {order[3]} на сумму {price[0][0] * order[3]}₽.\n"
                    cost += price[0][0] * order[3]
                    number += 1
            text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            text += f"\n\nДля начала оплаты заказа или возвращения товаров в корзину воспользуйтесь кнопками\U00002B07"
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Начать оплату", callback_data=f"pay_begin{num}")], [InlineKeyboardButton("\U0001F519\U0001F6D2Вернуть товары в корзину", callback_data=f"back_to_basket{num}")]])
            await app.send_message(chat_id=orders[0][4], text=text, reply_markup=markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0001F4E8Заказ №{num} частично подтверждён\U00002705", message_id=function_call.message.id)
        else:
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Пользователь отменил заказ или забронированный товар был удалён из списка товаров", message_id=function_call.message.id)
        connection.close()
    elif "back_to_basket" in function_call.data:
        num = (function_call.data)[14:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT product, count FROM orders WHERE num = ?", (num,))
        products = cursor.fetchall()
        for product in products:
            cursor.execute("SELECT num, count FROM orders WHERE customer_id = ? and status = ? and product = ?", (function_call.from_user.id, 1, product[0]))
            count = cursor.fetchall()
            if count != []:
                cursor.execute("UPDATE orders SET count = ? WHERE num = ?", (count[0][1] + product[1], count[0][0]))
                cursor.execute("DELETE FROM orders WHERE num = ? and product = ?", (num, product[0]))
                connection.commit()
        cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (1, num))
        connection.commit()
        cursor.execute("SELECT customer_id FROM orders WHERE num = ?", (num,))
        cust_id = cursor.fetchall()
        cursor.execute("UPDATE orders SET num = ? WHERE status = ? and customer_id = ?", (num, 1, cust_id[0][0]))
        connection.commit()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 1))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            for order in orders:
                if num < 5:
                    cursor.execute("SELECT count, price FROM Plants WHERE num = ?", (order[1], ))
                    osts = cursor.fetchall()
                    button = [InlineKeyboardButton(text=f"{order[2]}, {order[3]}шт. {order[3] * osts[0][1]}₽", callback_data=f"info_product{order[0]}num{order[1]}")]
                    num += 1
                    markup.append(button)
            if len(orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"next_basket5")])
            markup.append([InlineKeyboardButton(text="\U00002705Оформить заказ\U00002705", callback_data="new_order")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0000203CПри заказе товаров из разных закупок доставка будет осуществляться несколькими отправлениями\n\nДля оформления заказа воспользуйтесь кнопокой\U0001F447", reply_markup=markup, message_id=function_call.message.id)
        else:
            await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
            await app.send_message(chat_id=function_call.from_user.id, text="\U0001F6D2**Корзина**\n\n\U0001F5D1Ваша корзина пуста. Добавьте товары через комментарии!\U0001F4AC")
        connection.close()
    elif "pay_begin" in function_call.data:
        num = (function_call.data)[9:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        if orders != []:
            cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (3, num))
            connection.commit()
            text = f"\U0001F7E2\U0001F4B3Вы начали оплату заказа №`{num}`\n\n\U0001F4E6Состав заказа:\n"
            spis_purchages = []
            for order in orders:
                if order[5] not in spis_purchages:
                    spis_purchages.append(order[5])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                number = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                orders = cursor.fetchall()
                for order in orders:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (order[1],))
                    price = cursor.fetchall()
                    text += f"{number}. Товар {order[1]} **{order[2]}**. Кол-во: {order[3]} на сумму {price[0][0] * order[3]}₽.\n"
                    cost += price[0][0] * order[3]
                    number += 1
            text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            cursor.execute("SELECT card_data FROM Payment")
            card_data = cursor.fetchall()
            text += f"\n\n\U0001F4B3Оплатите заказ по реквизитам:\n{card_data[0][0]}"
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Я оплатил", callback_data=f"buy_done{num}")], [InlineKeyboardButton("\U0000274CОтказаться от оплаты", callback_data=f"buy_user_reject{num}")]])
            await app.edit_message_text(chat_id=orders[0][4], text=text, reply_markup=markup, message_id=function_call.message.id)
        else:
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Вы отменили заказ или забронированный товар был удалён из списка товаров", message_id=function_call.message.id)
        connection.close()
    elif "buy_user_reject" in function_call.data:
        num = (function_call.data)[15:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        if orders != []:
            text = f"\U0001F534\U0001F6D2Вы успешно отказались от оплаты заказа №`{num}`\n\n\U0000274CЗаказ удалён\n\n\U0001F4E6Состав заказа:\n"
            spis_purchages = []
            for order in orders:
                if order[5] not in spis_purchages:
                    spis_purchages.append(order[5])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                number = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                orderss = cursor.fetchall()
                for orderr in orderss:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (orderr[1],))
                    price = cursor.fetchall()
                    text += f"{number}. Товар {orderr[1]} **{orderr[2]}**. Кол-во: {orderr[3]} на сумму {price[0][0] * orderr[3]}₽.\n"
                    cost += price[0][0] * orderr[3]
                    number += 1
            text += f"\n\U0001F4B5Общая стоимость: {cost}₽"
            for order in orders:
                cursor.execute("SELECT book FROM Plants WHERE num = ?", (order[0],))
                book = cursor.fetchall()
                cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - order[2], order[0]))
            cursor.execute("DELETE FROM orders WHERE num = ?", (num,))
            connection.commit()
            await app.edit_message_text(chat_id=function_call.from_user.id, text=text, message_id=function_call.message.id)
        else:
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Вы уже отменили заказ или забронированный товар был удалён из списка товаров", message_id=function_call.message.id)
        connection.close()
    elif "buy_done" in function_call.data:
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F4CEПожалуйста, отпавьте скриншот чека об оплате заказа или сам чек в формате pdf", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, buy_done, kwargs={"num": int((function_call.data)[8:])})
    elif "buy_reject" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (int(function_call.data[10:]),))
        orders = cursor.fetchall()
        num = int(function_call.data[10:])
        if orders != []:
            text = f"\U0001F534\U0001F6D2Оплата заказа №`{int(function_call.data[10:])}` отклонена продавцом\n\n\U0001F4E6Состав заказа:\n"
            spis_purchages = []
            for order in orders:
                if order[6] not in spis_purchages:
                    spis_purchages.append(order[6])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                number = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                orderss = cursor.fetchall()
                for orderr in orderss:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (orderr[1],))
                    price = cursor.fetchall()
                    text += f"{number}. Товар {orderr[1]} **{orderr[2]}**. Кол-во: {orderr[3]} на сумму {price[0][0] * orderr[3]}₽.\n"
                    cost += price[0][0] * orderr[3]
                    number += 1
            text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
            for order in orders:
                cursor.execute("SELECT book FROM Plants WHERE num = ?", (order[1],))
                count = cursor.fetchall()
                cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (count[0][0] - order[3], order[1]))
            connection.commit()
            await app.send_message(chat_id=orders[0][4], text=text)
            cursor.execute("DELETE FROM orders WHERE num = ?", (int(function_call.data[10:]),))
            connection.commit()
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0001F534\U00002705Оплата заказа №`{orders[0][0]}` отклонена успешно.\n\n\U0001F464Связаться с клиентом: {orders[0][5]}", message_id=function_call.message.id)
        else:
            await app.send_message(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Вы отменили заказ или забронированный товар был удалён из списка товаров")
        connection.close()
    elif "buy_confirm" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name FROM orders WHERE num = ?", (int(function_call.data[11:]),))
        order = cursor.fetchall()
        for ord in order:
            cursor.execute("SELECT ordered FROM Plants WHERE num = ?", (ord[1],))
            count = cursor.fetchall()
            cursor.execute("UPDATE Plants SET ordered = ? WHERE num = ?", (count[0][0] + ord[3], ord[1]))
        connection.commit()
        if order != []: 
            cursor.execute("UPDATE orders SET is_pay = ?, status = ? WHERE num = ?", (True, 5, int(function_call.data[11:])))
            connection.commit()
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0001F7E2\U00002705Оплата заказа №`{order[0][0]}` успешно подтверждена. Данные для доставки запрошены\n\n\U0001F464Связаться с клиентом: {order[0][5]}", message_id=function_call.message.id)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("Почта России", callback_data=f"send_post{int(function_call.data[11:])}")], [InlineKeyboardButton("СДЭК", callback_data=f"send_sdek{int(function_call.data[11:])}")], [InlineKeyboardButton("Самовывоз", callback_data=f"send_sam{int(function_call.data[11:])}")]])
            await app.send_message(chat_id=order[0][4], text=f"\U00002714Оплата заказа №`{order[0][0]}` подтреждена продавцом.\n\n\U00002709Пожалуйста, воспользуйтесь кнопками для выбора способа получения заказа", reply_markup=markup)
        else:
            await app.send_message(chat_id=function_call.from_user.id, text="\U0000274CОшибка. Вы отменили заказ или забронированный товар был удалён из списка товаров")
        connection.close()
    elif "send_post" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("UPDATE orders SET is_post = ? WHERE num = ?", ((function_call.data)[5:9], int(function_call.data[9:])))
        connection.commit()
        if "send_post" in function_call.data:
            send = "Почта России"
        elif "send_sdek" in function_call.data:
            send = "СДЭК"
        cursor.execute("SELECT id, name, adress, telephone from Users WHERE id = ?", (function_call.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("Ускоренная доставка", callback_data=f"first_post{int(function_call.data[9:])}")], [InlineKeyboardButton("Обычная доставка", callback_data=f"secnd_post{int(function_call.data[9:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_back_type{int(function_call.data[9:])}")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали способ доставки: **{send}**\n\nТеперь выберите тип доставки заказа\U0001F4EC", message_id=function_call.message.id, reply_markup=markup)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0000270FВвести вручную", callback_data=f"send_name{int(function_call.data[9:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_back_type{int(function_call.data[9:])}")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали способ доставки: **{send}**\n\n\U0001F4CDВоспользуйтесь кнопкой ниже, чтобы ввести данные для доставки\n\n\U0001F4DDЧтобы каждый раз не вводить данные вручную, заполните их с помощью команды /account\U00002B07", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif "first_post" in function_call.data or "secnd_post" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        if "first_post" in function_call.data:
            word_send = "ускоренную доставку"
            type_send = 1
        elif "secnd_post" in function_call.data:
            type_send = 0
            word_send = "обычную доставку"
        cursor.execute("UPDATE orders SET class = ? WHERE num = ?", (type_send, int(function_call.data[10:])))
        connection.commit()
        cursor.execute("SELECT id, name, adress, telephone from Users WHERE id = ?", (function_call.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F464Взять из личного кабинета", callback_data=f"use_data{int(function_call.data[10:])}")], [InlineKeyboardButton("\U0000270FВвести вручную", callback_data=f"send_name{int(function_call.data[10:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_post{int(function_call.data[10:])}")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали **{word_send}**\n\n\U0001F4CDВоспользуйтесь кнопками ниже, чтобы взять данные для доставки из личного кабинета или ввести их вручную\U00002B07", message_id=function_call.message.id, reply_markup=markup)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0000270FВвести вручную", callback_data=f"send_name{int(function_call.data[10:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_back_type{int(function_call.data[10:])}")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали способ доставки: **{send}**\n\n\U0001F4CDВоспользуйтесь кнопкой ниже, чтобы ввести данные для доставки\n\n\U0001F4DDЧтобы каждый раз не вводить данные вручную, заполните их с помощью команды /account\U00002B07", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif "send_sdek" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("UPDATE orders SET is_post = ? WHERE num = ?", ((function_call.data)[5:9], int(function_call.data[9:])))
        connection.commit()
        if "send_post" in function_call.data:
            send = "Почта России"
        elif "send_sdek" in function_call.data:
            send = "СДЭК"
        cursor.execute("SELECT id, name, adress, telephone from Users WHERE id = ?", (function_call.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F464Взять из личного кабинета", callback_data=f"use_data{int(function_call.data[9:])}")], [InlineKeyboardButton("\U0000270FВвести вручную", callback_data=f"send_name{int(function_call.data[9:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_back_type{int(function_call.data[9:])}")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали способ доставки: **{send}**\n\n\U0001F4CDВоспользуйтесь кнопками ниже, чтобы взять данные для доставки из личного кабинета или ввести их вручную\U00002B07", message_id=function_call.message.id, reply_markup=markup)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0000270FВвести вручную", callback_data=f"send_name{int(function_call.data[9:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_back_type{int(function_call.data[9:])}")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали способ доставки: **{send}**\n\n\U0001F4CDВоспользуйтесь кнопкой ниже, чтобы ввести данные для доставки\n\n\U0001F4DDЧтобы каждый раз не вводить данные вручную, заполните их с помощью команды /account\U00002B07", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif "send_sam" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("UPDATE orders SET is_post = ? WHERE num = ?", ("sam", int(function_call.data[8:])))
        connection.commit()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Отправить данные", callback_data=f"send_data{int(function_call.data[8:])}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_back_type{int(function_call.data[8:])}")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"Вы выбрали способ доставки: **Самовывоз**", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "use_data" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (function_call.from_user.id,))
        user = cursor.fetchall()
        cursor.execute("UPDATE orders SET name = ?, adress = ?, telephone = ? WHERE num = ?", (user[0][1], user[0][2], user[0][3], int(function_call.data[8:])))
        connection.commit()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, class FROM orders WHERE num = ?", (int(function_call.data[8:]),))
        order = cursor.fetchall() 
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U00002705Данные по заказу №`{order[0][0]}` переданы продавцу. Он свяжется с вами в ближайшее время", message_id=function_call.message.id)
        cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (6, int(function_call.data[8:])))
        connection.commit()
        if order[0][6] == "post":
            send = "Почта России"
        elif order[0][6] == "sdek":
            send = "СДЭК"
        else:
            send = "Самовывоз"
        if order[0][6] == "post":
            if order[0][10] == 1:
                word_send = "**Тип доставки**: Ускоренная доставка\n"
            else:
                word_send = "**Тип доставки**: Обычная доставка\n"
        else:
            word_send = ""
        if send != "Самовывоз":
            await app.send_message(chat_id=993684230, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}\n{word_send}**ФИО**: {order[0][7]}\n**Адрес доставки**: {order[0][8]}\n**Телефон**: {order[0][9]}")
            await app.send_message(chat_id=1333967466, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}\n{word_send}**ФИО**: {order[0][7]}\n**Адрес доставки**: {order[0][8]}\n**Телефон**: {order[0][9]}")
        else:
            await app.send_message(chat_id=1333967466, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}")
            await app.send_message(chat_id=993684230, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}")
        connection.close()
    elif "send_back_type" in function_call.data:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("Почта России", callback_data=f"send_post{int(function_call.data[14:])}")], [InlineKeyboardButton("СДЭК", callback_data=f"send_sdek{int(function_call.data[14:])}")], [InlineKeyboardButton("Самовывоз", callback_data=f"send_sam{int(function_call.data[14:])}")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U00002709Пожалуйста, воспользуйтесь кнопками для выбора способа получения заказа", message_id=function_call.message.id, reply_markup=markup)
    elif "send_name" in function_call.data:
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FВведите ваше ФИО", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, send_name, kwargs={"num": int((function_call.data)[9:])})
    elif "send_adress" in function_call.data:
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FВведите адрес доставки", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, send_adress, kwargs={"num": int((function_call.data)[11:])})
    elif "send_phone" in function_call.data:
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F4CEОтправить мой номер", request_contact=True)]], resize_keyboard=True)
        await app.send_message(chat_id=function_call.from_user.id, text="Введите ваш номер телефона или воспользуйтесь кнопкой\U00002B07", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, send_phone, kwargs={"num": int((function_call.data)[10:])})
    elif "send_data" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, class FROM orders WHERE num = ?", (int(function_call.data[9:]),))
        order = cursor.fetchall() 
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U00002705Данные по заказу №`{order[0][0]}` переданы продавцу. Он свяжется с вами в ближайшее время", message_id=function_call.message.id)
        cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (6, int(function_call.data[9:])))
        connection.commit()
        if order[0][6] == "post":
            send = "Почта России"
        elif order[0][6] == "sdek":
            send = "СДЭК"
        else:
            send = "Самовывоз"
        if order[0][6] == "post":
            if order[0][10] == 1:
                word_send = "**Тип доставки**: Ускоренная доставка\n"
            else:
                word_send = "**Тип доставки**: Обычная доставка\n"
        else:
            word_send = ""
        if send != "Самовывоз":
            await app.send_message(chat_id=1333967466, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}\n{word_send}**ФИО**: {order[0][7]}\n**Адрес доставки**: {order[0][8]}\n**Телефон**: {order[0][9]}")
            await app.send_message(chat_id=993684230, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}\n{word_send}**ФИО**: {order[0][7]}\n**Адрес доставки**: {order[0][8]}\n**Телефон**: {order[0][9]}")
        else:
            await app.send_message(chat_id=1333967466, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}")
            await app.send_message(chat_id=993684230, text=f"\U0001F4E4Покупатель {order[0][5]} прислал данные для отправки заказа №`{order[0][0]}`\n\n**Способ доставки**: {send}")
        connection.close()
    elif function_call.data == "change_recieve_order":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and is_pay = ?", (function_call.from_user.id, 1))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            spis_orders = []
            for order in orders:
                if order[0] not in spis_orders:
                    spis_orders.append(order[0])
            spis_orders = sorted(spis_orders)
            for ord in spis_orders:
                if num < 5:
                    button = [InlineKeyboardButton(text=f"Заказ №{ord}", callback_data=f"change_order{ord}")]
                    num += 1
                    markup.append(button)
            if len(spis_orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"change_next_order5")])
            markup.append([InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FВыберите заказ, данные для получения которого нужно изменить\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FИзменить данные для получения можно только для неоплаченного заказа\n\nНеоплаченные заказы отсутствуют", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "change_next_order" in function_call.data:
        next_b = (function_call.data)[17:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and is_pay = ?", (function_call.from_user.id, 1))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            spis_orders = []
            for order in orders:
                if order[0] not in spis_orders:
                    spis_orders.append(order[0])
            spis_orders = sorted(spis_orders)
            for ord in spis_orders:
                if num >= int(next_b) and num < int(next_b) + 5:
                    button = [InlineKeyboardButton(text=f"Заказ №{ord}", callback_data=f"change_order{ord}")]
                    markup.append(button)
                num += 1
            if int(next_b) >= 5:
                markup.append([InlineKeyboardButton(text="Назад\U000025C0", callback_data=f"change_next_order{int(next_b) - 5}")])
            if len(spis_orders) - int(next_b) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"change_next_order{int(next_b) + 5}")])
            markup.append([InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FВыберите заказ, данные для получения которого нужно изменить\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0000270FИзменить данные для получения можно только для неоплаченного заказа\n\nНеоплаченные заказы отсутствуют", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif "change_order" in function_call.data:
        num = (function_call.data)[12:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, reciept, is_pay FROM orders WHERE num = ?", (num,))
        orders = cursor.fetchall()
        cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (5, num))
        connection.commit()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("Почта России", callback_data=f"send_post{num}")], [InlineKeyboardButton("СДЭК", callback_data=f"send_sdek{num}")], [InlineKeyboardButton("Самовывоз", callback_data=f"send_sam{num}")]])
        await app.send_message(chat_id=function_call.from_user.id, text=f"\U00002709Воспользуйтесь кнопками для выбора способа получения заказа.\n\n\U0000203CПри доставке почтой оплата производится по тарифам почты России. Оплата после отправки по тем же реквизитам. При отправке СДЭКом доставка оплачивается при получении заказа. Упаковка для всех отправок (кроме самовывоза) 100₽", reply_markup=markup)
        connection.close()
    elif function_call.data == "pay_my_order":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 3))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            spis_orders = []
            for order in orders:
                if order[0] not in spis_orders:
                    spis_orders.append(order[0])
            spis_orders = sorted(spis_orders)
            for ord in spis_orders:
                if num < 5:
                    button = [InlineKeyboardButton(text=f"Заказ №{ord}", callback_data=f"pay_begin{ord}")]
                    num += 1
                    markup.append(button)
            if len(spis_orders) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"do_pay_next_order5")])
            markup.append([InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F4B5Выберите заказ, который хотите оплатить\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F4B5Оплатить можно только подтверждённый продавцом заказ\n\nПодтверждённые продавцом заказы отсутствуют", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "do_pay_next_order" in function_call.data:
        next_b = (function_call.data)[17:]
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, is_pay, status FROM orders WHERE customer_id = ? and status = ?", (function_call.from_user.id, 3))
        orders = cursor.fetchall()
        if orders != []:
            markup = []
            num = 0
            spis_orders = []
            for order in orders:
                if order[0] not in spis_orders:
                    spis_orders.append(order[0])
            spis_orders = sorted(spis_orders)
            for ord in spis_orders:
                if num >= int(next_b) and num < int(next_b) + 5:
                    button = [InlineKeyboardButton(text=f"Заказ №{ord}", callback_data=f"pay_begin{ord}")]
                    markup.append(button)
                num += 1
            if int(next_b) >= 5:
                markup.append([InlineKeyboardButton(text="Назад\U000025C0", callback_data=f"do_pay_next_order{int(next_b) - 5}")])
            if len(spis_orders) - int(next_b) > 5:
                markup.append([InlineKeyboardButton(text="Далее\U000025B6", callback_data=f"do_pay_next_order{int(next_b) + 5}")])
            markup.append([InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")])
            markup = InlineKeyboardMarkup(markup)
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F4B5Выберите заказ, который хотите оплатить\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F4B5Оплатить можно только подтверждённый продавцом заказ\n\nПодтверждённые продавцом заказы отсутствуют", message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif function_call.data == "delivery_time" or function_call.data == "delivery_time_account":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT time1, time2, time3, time4, time5, name FROM purchages")
        purchages = cursor.fetchall()
        time_text = ""
        for purchage in purchages:
            time_text += f"**Закупка {purchage[5]}**\n\U0001F4CD**Обработка заказа:** {purchage[0]}\n\U0001F4CD**Отправка заказа садовнику:** {purchage[1]}\n\U0001F4CD**Оплата заказа:** {purchage[2]}\n\U0001F4CD**Поступление растений:** {purchage[3]}\n\U0001F4CD**Отправка растений:** {purchage[4]}\n\n"
        connection.close()
        if function_call.data == "delivery_time":
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"basket")]])
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F519Назад", callback_data=f"my_orders")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text=time_text, reply_markup=markup, message_id=function_call.message.id)
    elif "ord_tble_purchage" in function_call.data:
        name_purchage = (function_call.data)[17:]
        await app.edit_message_text(chat_id=function_call.from_user.id, text=f"\U0001F551Создаю таблицу для закупки **{name_purchage}**", message_id=function_call.message.id)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, name FROM Plants")
        plants = cursor.fetchall()
        head = ["Номер заказа", "Закупка", "Оплата", "Статус", "ID заказчика", "Контакт заказчика", "Доставка", "Тип доставки", "ФИО", "Адрес доставки", "Телефон"]
        plants_nums = []
        for plant in plants:
            head.append(plant[1])
            plants_nums.append(plant[0])
        cursor.execute("SELECT num FROM orders WHERE purchage = ? ORDER BY num", (name_purchage,))
        nums = cursor.fetchall()
        spis_nums = []
        data = []
        spis_ids = []
        for num in nums:
            if num[0] not in spis_nums:
                spis_nums.append(num[0])
                cursor.execute("SELECT * FROM orders WHERE num = ? and purchage = ?", (num[0], name_purchage))
                orders = cursor.fetchall()
                if orders[0][5] not in spis_ids:
                    spis_ids.append(orders[0][5])
                    datapr = []
                    cursor.execute("SELECT * FROM orders WHERE num = ? ORDER BY purchage", (num[0],))
                    spis_num_need = cursor.fetchall()
                    need_num = 1
                    for spis_need_num_one in spis_num_need:
                        if spis_need_num_one[1] == name_purchage:
                            break
                        need_num += 1
                    flag = False
                    for spis_need_num_one in spis_num_need:
                        if spis_need_num_one[1] != name_purchage:
                            flag = True
                    if flag == True:
                        datapr.append(str(orders[0][0]) + "." + str(need_num))
                    else:
                        datapr.append(str(orders[0][0]))
                    datapr.append(name_purchage)
                    datapr.append(orders[0][7])
                    if orders[0][13] == 1:
                        datapr.append("В корзине")
                    elif orders[0][13] == 2:
                        datapr.append("Ожидает проверки наличия товара")
                    elif orders[0][13] == 3:
                        datapr.append("Ожидает оплаты")
                    elif orders[0][13] == 4:
                        datapr.append("Ожидает проверки оплаты")
                    elif orders[0][13] == 5:
                        datapr.append("Ожидает ввода данных для отправки")
                    elif orders[0][13] == 6:
                        datapr.append("Оплата подтверждена, заказ принят")
                    datapr.append(orders[0][5])
                    datapr.append(orders[0][6])
                    datapr.append(orders[0][8])
                    if orders[0][7] == "post":
                        if orders[0][13] == 1:
                            word_send = "Ускоренная"
                        else:
                            word_send = "Обычная"
                    else:
                        word_send = "-"
                    datapr.append(word_send)
                    datapr.append(orders[0][9])
                    datapr.append(orders[0][10])
                    datapr.append(orders[0][11])
                    product = {}
                    for order in orders:
                        product[f"{order[2]}"] = order[4]
                    for plant in plants_nums:
                        count = product.get(f"{plant}")
                        datapr.append(count)
                    cursor.execute("SELECT * FROM orders WHERE customer_id = ? and purchage = ?", (orders[0][5], name_purchage))
                    orders_id = cursor.fetchall()
                    data.append(datapr)
                    spis_nums_id = []
                    for order_id in orders_id:
                        if order_id[0] != num[0] and order_id[0] not in spis_nums_id:
                            spis_nums_id.append(order_id[0])
                            cursor.execute("SELECT * FROM orders WHERE num = ? and purchage = ?", (order_id[0], name_purchage))
                            orders_one_id = cursor.fetchall()
                            datapr = []
                            cursor.execute("SELECT * FROM orders WHERE num = ? ORDER BY purchage", (order_id[0],))
                            spis_num_need = cursor.fetchall()
                            need_num = 1
                            for spis_need_num_one in spis_num_need:
                                if spis_need_num_one[1] == name_purchage:
                                    break
                                need_num += 1
                            flag = False
                            for spis_need_num_one in spis_num_need:
                                if spis_need_num_one[1] != name_purchage:
                                    flag = True
                            if flag == True:
                                datapr.append(str(orders_one_id[0][0]) + "." + str(need_num))
                            else:
                                datapr.append(str(orders_one_id[0][0]))
                            datapr.append(name_purchage)
                            datapr.append(orders_one_id[0][7])
                            if orders_one_id[0][13] == 1:
                                datapr.append("В корзине")
                            elif orders_one_id[0][13] == 2:
                                datapr.append("Ожидает проверки наличия товара")
                            elif orders_one_id[0][13] == 3:
                                datapr.append("Ожидает оплаты")
                            elif orders_one_id[0][13] == 4:
                                datapr.append("Ожидает проверки оплаты")
                            elif orders_one_id[0][13] == 5:
                                datapr.append("Ожидает ввода данных для отправки")
                            elif orders_one_id[0][13] == 6:
                                datapr.append("Оплата подтверждена, заказ принят")
                            datapr.append(orders_one_id[0][5])
                            datapr.append(orders_one_id[0][6])
                            datapr.append(orders_one_id[0][8])
                            if orders_one_id[0][7] == "post":
                                if orders_one_id[0][13] == 1:
                                    word_send = "Ускоренная"
                                else:
                                    word_send = "Обычная"
                            else:
                                word_send = "-"
                            datapr.append(word_send)
                            datapr.append(orders_one_id[0][9])
                            datapr.append(orders_one_id[0][10])
                            datapr.append(orders_one_id[0][11])
                            product = {}
                            for order_one in orders_one_id:
                                product[f"{order_one[2]}"] = order_one[4]
                            for plant in plants_nums:
                                count = product.get(f"{plant}")
                                datapr.append(count)
                            data.append(datapr)
        wb = Workbook()
        ws = wb.active
        ws.append(head)
        for i in range(len(data)):
            ws.append(data[i])
        for i in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            ws[f'{i}1'].font = Font(bold=True)
        ws[f'AA1'].font = Font(bold=True)
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 50
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 20
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 20
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 20
        ws.column_dimensions['T'].width = 20
        ws.column_dimensions['U'].width = 20
        ws.column_dimensions['V'].width = 20
        ws.column_dimensions['W'].width = 20
        ws.column_dimensions['X'].width = 20
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 20
        ws.column_dimensions['AA'].width = 20
        ws.column_dimensions['AB'].width = 20
        ws.column_dimensions['AC'].width = 20
        ws.column_dimensions['AD'].width = 20
        ws.column_dimensions['AE'].width = 20
        ws.column_dimensions['AF'].width = 20
        ws.column_dimensions['AG'].width = 20
        for col in ws.columns:
            for cell in col:
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj
        connection.close()
        wb.save(f"C:/Users/Administrator/My_bots/Botanica_club_bot/{name_purchage}.xlsx")
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
        await app.send_document(chat_id=function_call.from_user.id, document=f"C:/Users/Administrator/My_bots/Botanica_club_bot/{name_purchage}.xlsx", reply_markup=markup)
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
    elif function_call.data == "exit_table2":
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002699Управление заказами", callback_data=f"red_table_orders")], [InlineKeyboardButton("\U0001F69BУправление закупками", callback_data=f"admin_purchage")], [InlineKeyboardButton("\U0000270FРедактировать таблицу товаров", callback_data=f"admin_table")], [InlineKeyboardButton("\U0001F440Посмотреть таблицу товаров", callback_data=f"view_table")]])
        await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
        await app.send_message(chat_id=function_call.from_user.id, text="\U00002699Для использования функций администратора воспользуйтесь кнопками ниже\U00002B07", reply_markup=markup)
    elif function_call.data == "add_admin_order":
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F7E2\U0001F195Введите информацию о получателе заказа в формате **Телефон - Юзернейм в Telegram - Способ получения - ФИО - Адрес доставки**\n\n\U00002705Пример: +79163456789 - @example - Почта России - Иванов Иван Иванович - г. Примерный, ул. Примерная, д. 1, кв. 1", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, add_admin_order)
    elif "new_product" in function_call.data:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT telephone, customer_name, is_post, name, adress FROM orders WHERE num = ?", ((function_call.data)[11:],))
        order = cursor.fetchall()
        connection.close()
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U00002795Добавьте в заказ товар в формате: **Номер товара - Количество товара**\n\n\U00002705Пример: 2 - 4", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, new_product, kwargs={"telephone": order[0][0], "username": order[0][1], "is_post": order[0][2], "name": order[0][3], "adress": order[0][4]})
    elif "save_admin_order" in function_call.data:
        num = (function_call.data)[16:]
        num = int(num)
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("UPDATE orders SET status = ? WHERE num = ?", (6, num))
        connection.commit()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, reciept, is_pay, purchage FROM orders WHERE num = ?", (num,))
        order = cursor.fetchall()
        if order != []:
            send = order[0][6]
            if order[0][11] == True:
                money = "Оплачен"
            else:
                money = "Не оплачен"
            if order[0][11] == True:
                text = f"\U00002705Заказ сохранён!\n\n**Номер заказа:** {order[0][0]}\n**Заказчик:** {order[0][5]}\n**Оплата:** {money}\n**Способ доставки:** {send}\n**ФИО закзачика:** {order[0][7]}\n**Адрес доставки:** {order[0][8]}\n**Телефон:** {order[0][9]}\n\n**Список товаров:**\n"
                spis_purchages = []
                for ord in order:
                    if ord[12] not in spis_purchages:
                        spis_purchages.append(ord[12])
                cost = 0
                spis_purchages = sorted(spis_purchages)
                for purchage in spis_purchages:
                    text += f"**Закупка {purchage}**\n"
                    number = 1
                    cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                    orderss = cursor.fetchall()
                    for ordd in orderss:
                        cursor.execute("SELECT price FROM Plants WHERE num = ?", (ordd[1],))
                        price = cursor.fetchall()
                        text += f"{number}. Товар {ordd[1]} **{ordd[2]}**. Кол-во: {ordd[3]} на сумму {price[0][0] * ordd[3]}₽.\n"
                        cost += price[0][0] * ordd[3]
                        number += 1
                text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
                text = text.replace("None", "Не указано")
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.edit_message_text(chat_id=function_call.from_user.id, text=text, message_id=function_call.message.id, reply_markup=markup)
        connection.close()
    elif function_call.data == "view_order":
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F440Введите номер заказа для просмотра информации о нём", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, view_order)
    elif function_call.data == "del_order":
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U00002716Введите номер заказа для его удаления", message_id=function_call.message.id)
        await pyrostep.register_next_step(function_call.from_user.id, del_order)
    elif function_call.data == "delete_unpay":
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num FROM orders")
        nums = cursor.fetchall()
        spis_nums = []
        for num in nums:
            if num[0] not in spis_nums:
                spis_nums.append(num[0])
                cursor.execute("SELECT num, product, product_name, customer_id, is_pay, count, purchage FROM orders WHERE num = ?", (num[0],))
                orders = cursor.fetchall()
                if orders[0][4] == None:
                    num = 1
                    text=f"\U0001F534\U0001F6D2Ваш заказ №`{orders[0][0]}` удалён.\nПричина: \U0000274Cсрок оплаты истёк\n\n\U0001F4E6Состав заказа:\n"
                    spis_purchages = []
                    cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ?", (orders[0][0],))
                    order = cursor.fetchall()
                    for ord in order:
                        if ord[8] not in spis_purchages:
                            spis_purchages.append(ord[8])
                    cost = 0
                    spis_purchages = sorted(spis_purchages)
                    for purchage in spis_purchages:
                        text += f"**Закупка {purchage}**\n"
                        number = 1
                        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (orders[0][0], purchage))
                        orderss = cursor.fetchall()
                        for ordd in orderss:
                            cursor.execute("SELECT price FROM Plants WHERE num = ?", (ordd[1],))
                            price = cursor.fetchall()
                            text += f"{number}. Товар {ordd[1]} **{ordd[2]}**. Кол-во: {ordd[3]} на сумму {price[0][0] * ordd[3]}₽.\n"
                            cost += price[0][0] * ordd[3]
                            number += 1
                    text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
                    await app.send_message(chat_id=orders[0][3], text=text)
                    cursor.execute("DELETE FROM orders WHERE num = ?", (orders[0][0],))
                    connection.commit()
                    cursor.execute("SELECT book FROM Plants WHERE num = ?", (orders[0][1],))
                    book = cursor.fetchall()
                    cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] - orders[0][5], orders[0][1]))
                    connection.commit()
            else:
                continue
        connection.close()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.edit_message_text(chat_id=function_call.from_user.id, text="\U0001F534\U0001F4B8Неоплаченные заказы удалены", message_id=function_call.message.id, reply_markup=markup)

async def del_order(client, message):
    if (message.text).isdigit():
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, reciept, is_pay, purchage FROM orders WHERE num = ?", (int(message.text),))
        order = cursor.fetchall()
        if order != []:
            text = f"\U0001F534\U0001F6D2Продавец удалил ваш заказ №`{order[0][0]}`\n\n\U0001F4E6Состав заказа:\n"
            spis_purchages = []
            for ord in order:
                if ord[12] not in spis_purchages:
                    spis_purchages.append(ord[12])
            cost = 0
            spis_purchages = sorted(spis_purchages)
            for purchage in spis_purchages:
                text += f"**Закупка {purchage}**\n"
                number = 1
                cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (int(message.text), purchage))
                orderss = cursor.fetchall()
                for ordd in orderss:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (ordd[1],))
                    price = cursor.fetchall()
                    text += f"{number}. Товар {ordd[1]} **{ordd[2]}**. Кол-во: {ordd[3]} на сумму {price[0][0] * ordd[3]}₽.\n"
                    cost += price[0][0] * ordd[3]
                    number += 1
            text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
            try:
                await app.send_message(chat_id=order[0][4], text=text)
            except:
                pass
            for ord in order:
                cursor.execute("SELECT book, ordered FROM Plants WHERE num = ?", (ord[1],))
                count = cursor.fetchall()
                cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (count[0][0] - ord[3], ord[1]))
                if ord[11] == 1:
                    cursor.execute("UPDATE Plants SET ordered = ? WHERE num = ?", (count[0][1] - ord[3], ord[1]))
                connection.commit()
            cursor.execute("DELETE FROM orders WHERE num = ?", (int(message.text),))
            connection.commit()
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.send_message(chat_id=message.from_user.id, text=f"\U00002705Заказ №`{order[0][0]}` успешно удалён", reply_markup=markup)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет заказа с таким номером", reply_markup=markup)
        connection.close()
    else:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CНомер заказа введён неверно", reply_markup=markup)

async def view_order(client, message):
    if (message.text).isdigit():
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_post, name, adress, telephone, reciept, is_pay, class, purchage FROM orders WHERE num = ?", (int(message.text),))
        order = cursor.fetchall()
        if order != []:
            if order[0][6] == "post":
                send = "Почта России"
            elif order[0][6] == "sdek":
                send = "СДЭК"
            elif order[0][6] == "sam":
                send = "Самовывоз"
            else:
                send = order[0][6]
            if order[0][10] != None:
                file = "Прикреплён"
            else:
                file = "Отсутствует"
            if order[0][11] == True:
                money = "Оплачен"
            else:
                money = "Не оплачен"
            if order[0][6] == "post":
                if order[0][12] == 1:
                    word_send = "**Тип доставки**: Ускоренная доставка\n"
                else:
                    word_send = "**Тип доставки**: Обычная доставка\n"
            else:
                word_send = ""
            if order[0][11] == True:
                text = f"**Номер заказа:** {order[0][0]}\n**Заказчик:** {order[0][5]}\n**Оплата:** {money}\n**Способ доставки:** {send}\n{word_send}**ФИО закзачика:** {order[0][7]}\n**Адрес доставки:** {order[0][8]}\n**Телефон:** {order[0][9]}\n**Чек:** {file}\n\n**Список товаров:**\n"
                spis_purchages = []
                for ord in order:
                    if ord[13] not in spis_purchages:
                        spis_purchages.append(ord[13])
                cost = 0
                spis_purchages = sorted(spis_purchages)
                for purchage in spis_purchages:
                    text += f"**Закупка {purchage}**\n"
                    number = 1
                    cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (int(message.text), purchage))
                    orderss = cursor.fetchall()
                    for ordd in orderss:
                        cursor.execute("SELECT price FROM Plants WHERE num = ?", (ordd[1],))
                        price = cursor.fetchall()
                        text += f"{number}. Товар {ordd[1]} **{ordd[2]}**. Кол-во: {ordd[3]} на сумму {price[0][0] * ordd[3]}₽.\n"
                        cost += price[0][0] * ordd[3]
                        number += 1
                text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
                text = text.replace("None", "Не указано")
                if order[0][10] != None:
                    try:
                        if "jpg" in order[0][10]:
                            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
                            await app.send_photo(chat_id=message.from_user.id, photo=order[0][10], caption=text, reply_markup=markup)
                        else:
                            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
                            await app.send_document(chat_id=message.from_user.id, document=order[0][10], caption=text, reply_markup=markup)
                    except:
                        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
                        await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)  
                else:
                    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
                    await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
            else:
                text = f"**Номер заказа:** {order[0][0]}\n**Заказчик:** {order[0][5]}\n**Оплата:**{money}"
                num = 1
                for ord in order:
                    cursor.execute("SELECT price FROM Plants WHERE num = ?", (ord[1],))
                    price = cursor.fetchall()
                    text += f"{num}. **Товар:** {ord[2]}, **Количество:** {ord[3]} шт., **Стоимость:** {price[0][0] * ord[3]}₽\n"
                    num += 1
                text = text.replace("None", "Не указано")
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
                await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет заказа с таким номером", reply_markup=markup)
        connection.close()
    else:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table2")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CНомер заказа введён неверно", reply_markup=markup)

async def add_admin_order(client, message):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    telephone = ""
    username = ""
    is_post = ""
    name = ""
    adress = ""
    text = message.text
    try:
        flag = True
        index = text.find("-")
        if index == -1:
            flag = False
        telephone = text[:(index-1)]
        text = text[(index+2):]
        index = text.find("-")
        if index == -1:
            flag = False
        username = text[:(index-1)]
        text = text[(index+2):]
        index = text.find("-")
        if index == -1:
            flag = False
        is_post = text[:(index-1)]
        text = text[(index+2):]
        index = text.find("-")
        if index == -1:
            flag = False
        name = text[:(index-1)]
        text = text[(index+2):]
        adress = text
        if flag == False:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"add_admin_order")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка считывания данных. Проверьте пробелы и корректность введённых данных", reply_markup=markup)
            return
        markup = ReplyKeyboardMarkup([[KeyboardButton("\U0001F519Назад")]], resize_keyboard=True)
        await app.send_message(chat_id=message.from_user.id, text="\U00002795Добавьте в заказ товар в формате: **Номер товара - Количество товара**\n\n\U00002705Пример: 2 - 4", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, new_product, kwargs={"telephone": telephone, "username": username, "is_post": is_post, "name": name, "adress": adress})
    except:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"add_admin_order")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка считывания данных. Проверьте пробелы и корректность введённых данных", reply_markup=markup)
    connection.close()

async def new_product(client, message, telephone: str = True, username: str = True, is_post: str = True, name: str = True, adress: str = True):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U0001F7E2\U0001F195Введите информацию о получателе заказа в формате **Телефон - Юзернейм в Telegram - Способ получения - ФИО - Адрес доставки**\n\n\U00002705Пример: +79163456789 - @example - Почта России - Иванов Иван Иванович - г. Примерный, ул. Примерная, д. 1, кв. 1", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, add_admin_order)
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    num = ""
    count = ""
    text = message.text
    try:
        index = text.find("-")
        num = text[:(index-1)]
        text = text[(index+2):]
        count = text
        cursor.execute("SELECT name FROM Plants WHERE num = ?", (int(num),))
        prod = cursor.fetchall()
        if prod == []:
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет товара с таким номером. Попробуйте ещё раз.\n\n\U00002795Добавьте в заказ товар в формате: **Номер товара - Количество товара**\n\n\U00002705Пример: 2 - 4", reply_markup=markup)
            await pyrostep.register_next_step(message.from_user.id, new_product, kwargs={"telephone": telephone, "username": username, "is_post": is_post, "name": name, "adress": adress})
            return
        cursor.execute('SELECT num FROM orders WHERE telephone = ? and status = ?', (telephone, 1))
        cust_id = cursor.fetchall()
        if cust_id == []:
            cursor.execute('SELECT num FROM orders')
            cust_id = cursor.fetchall()
            if cust_id == []:
                cust_id = 1
            else:
                cust_id = cust_id[-1][0]
                cust_id += 1
        else:
            cust_id = cust_id[0][0]
        cursor.execute("SELECT purchage FROM Plants WHERE num = ?", (int(num),))
        purchage = cursor.fetchall()
        cursor.execute("INSERT INTO orders (num, purchage, product, product_name, count, customer_name, is_pay, is_post, name, adress, telephone, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (cust_id, purchage[0][0], num, prod[0][0], int(count), username, True, is_post, name, adress, telephone, 1))
        connection.commit()
        cursor.execute("SELECT book, ordered FROM Plants WHERE num = ?", (int(num),))
        book = cursor.fetchall()
        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (book[0][0] + int(count), int(num)))
        cursor.execute("UPDATE Plants SET ordered = ? WHERE num = ?", (book[0][1] + int(count), int(num)))
        connection.commit()
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Данные о получателе сохранены", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton(text="\U00002795Добавить товар", callback_data=f"new_product{cust_id}")], [InlineKeyboardButton(text="\U00002705Сохранить заказ", callback_data=f"save_admin_order{cust_id}")]])
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Товар добавлен. Воспользуйтесь кнопками\U00002B07", reply_markup=markup)
    except:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка в введённых данных. Попробуйте ещё раз\n\n\U00002795Добавьте в заказ товар в формате: **Номер товара - Количество товара**\n\n\U00002705Пример: 2 - 4", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, new_product, kwargs={"telephone": telephone, "username": username, "is_post": is_post, "name": name, "adress": adress})
    connection.close()

async def part_new_purchage(function_call):
    markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
    await app.send_message(chat_id=function_call.from_user.id, text="\U0000270FВведите название закупки и её сроки в формате:\n\nНазвание закупки; Обработка заказа; Отправка заказа садовнику; Оплата заказа; Поступление растений; Отправка растений\n\n\U00002705Пример:\nзакупкароссия1; 1-2 дней; 3-5 дней; 7 дней; 5-7 дней; 8-10 часов", reply_markup=markup)
    await app.delete_messages(chat_id=function_call.from_user.id, message_ids=function_call.message.id)
    await pyrostep.register_next_step(function_call.from_user.id, new_purchage)

async def admin_purchage(message):
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F195Новая закупка", callback_data=f"new_purchage")], [InlineKeyboardButton("\U00002716Закрыть закупку", callback_data=f"del_purchage")], [InlineKeyboardButton("\U0000270Fзменить закупку", callback_data=f"edit_purchage")], [InlineKeyboardButton("\U0001F519Назад", callback_data=f"exit_table")]])
    await app.send_message(chat_id=message.from_user.id, text="\U0000270FЧто вы хотите сделать?", reply_markup=markup)

async def new_purchage(client, message):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Создание закупки отменено", reply_markup=markup)
        await admin_purchage(message)
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    text = message.text
    name = ""
    time1 = ""
    time2 = ""
    time3 = ""
    time4 = ""
    time5 = ""
    flag = True
    index = text.find(";")
    if index == -1:
        flag = False
    name = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time1 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time2 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time3 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time4 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index != -1:
        flag = False
    time5 = text
    if flag == False:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание закупки отменено", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_purchage")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
        return
    cursor.execute("SELECT * FROM purchages WHERE name = ?", (name,))
    name_is = cursor.fetchall()
    if name_is != []:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание закупки отменено", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_purchage")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CЗакупка с таким названием уже существует", reply_markup=markup)
        return  
    cursor.execute("INSERT INTO purchages (name, time1, time2, time3, time4, time5) VALUES (?, ?, ?, ?, ?, ?)", (name, time1, time2, time3, time4, time5))
    connection.commit()
    cursor.execute("SELECT name, time1, time2, time3, time4, time5 FROM purchages WHERE name = ?", (name,))
    purchage = cursor.fetchall()
    connection.close()
    markup = ReplyKeyboardRemove()
    await app.send_message(chat_id=message.from_user.id, text="\U00002714Данные успешно прочитаны", reply_markup=markup)
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0000270FВвести другие данные", callback_data=f"again_new_purchage{purchage[0][0]}")], [InlineKeyboardButton("\U00002705Сохранить", callback_data=f"save_purchage{purchage[0][0]}")]])
    await app.send_message(chat_id=message.from_user.id, text=f"\U00002139Информация о новой закупке:\n\n**Название:** {purchage[0][0]}\n**Обработка заказа:** {purchage[0][1]}\n**Отправка заказа садовнику:** {purchage[0][2]}\n**Оплата заказа:** {purchage[0][3]}\n**Поступление растений:** {purchage[0][4]}\n**Отправка растений:** {purchage[0][5]}", reply_markup=markup)

async def edit_purchage(client, message, name: str = True):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    text = message.text
    time1 = ""
    time2 = ""
    time3 = ""
    time4 = ""
    time5 = ""
    flag = True
    index = text.find(";")
    if index == -1:
        flag = False
    time1 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time2 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time3 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index == -1:
        flag = False
    time4 = text[:(index)]
    text = text[(index+2):]
    index = text.find(";")
    if index != -1:
        flag = False
    time5 = text
    if flag == False:
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"edit_purchage")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
        return
    cursor.execute("UPDATE purchages SET time1 = ?, time2 = ?, time3 = ?, time4 = ?, time5 = ? WHERE name = ?", (time1, time2, time3, time4, time5, name))
    connection.commit()
    cursor.execute("SELECT name, time1, time2, time3, time4, time5 FROM purchages WHERE name = ?", (name,))
    purchage = cursor.fetchall()
    connection.close()
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0000270FВвести другие данные", callback_data=f"edt_name_purchage{purchage[0][0]}")], [InlineKeyboardButton("\U00002705Сохранить", callback_data=f"save_purchage{purchage[0][0]}")]])
    await app.send_message(chat_id=message.from_user.id, text=f"\U00002139Изменённая информация о закупке:\n\n**Название:** {purchage[0][0]}\n**Обработка заказа:** {purchage[0][1]}\n**Отправка заказа садовнику:** {purchage[0][2]}\n**Оплата заказа:** {purchage[0][3]}\n**Поступление растений:** {purchage[0][4]}\n**Отправка растений:** {purchage[0][5]}", reply_markup=markup)

async def new_table(client, message, name: str = True):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Создание таблицы отменено", reply_markup=markup)
        await admin_table(message)
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    cursor.execute('DELETE FROM Temp_plants')
    connection.commit()
    num = ""
    name_product = ""
    kolvo = ""
    text = message.text
    flag = True
    try:
        if text.find("-") == -1:
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
            return
        while text.find("-") != -1:
            index = text.find("-")
            if index == -1:
                flag = False
            num = text[:(index-1)]
            text = text[(index+2):]
            index = text.find("-")
            if index == -1:
                flag = False
            name_product = text[:(index-1)]
            text = text[(index+2):]
            index = text.find("-")
            if index == -1:
                flag = False
            kolvo = text[:(index-1)]
            text = text[(index+2):]
            index = text.find(",")
            if index != -1:
                price = text[:(index)]
            else:
                price = text
            text = text[(index+2):]
            if flag == False:
                markup = ReplyKeyboardRemove()
                await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
                await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
                return
            cursor.execute("SELECT * FROM Plants WHERE num = ?", (int(num),))
            is_num = cursor.fetchall()
            if is_num != []:
                markup = ReplyKeyboardRemove()
                await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
                await app.send_message(chat_id=message.from_user.id, text="\U0000274CНомера товаров не должны повторяться (даже если товары находятся в разных закупках)", reply_markup=markup)
                return
            cursor.execute('INSERT INTO Temp_plants (num, purchage, name, count, price) VALUES (?, ?, ?, ?, ?)', (int(num), name, name_product, int(kolvo), int(price)))
        connection.commit()
    except:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
        return
    cursor.execute("SELECT num, name, count, price FROM Temp_plants")
    all_plants = cursor.fetchall()
    text = f"Теперь таблица для закупки **{name}** выглядит следующим образом:\n\n**Номер | Название растения | Количество | Цена**\n"
    for plant in all_plants:
        text += f"{plant[0]} | {plant[1]} | {plant[2]} | {plant[3]}р.\n"
    markup = ReplyKeyboardRemove()
    await app.send_message(chat_id=message.from_user.id, text="\U00002714Данные успешно прочитаны", reply_markup=markup)
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Всё верно", callback_data=f"right_new_table{name}")], [InlineKeyboardButton("\U0000274CЗаполнить заново", callback_data=f"new_table")]])
    await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
    connection.close()
    
async def admin_table(message):
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F195Новая таблица", callback_data=f"new_table")], [InlineKeyboardButton("\U00002796Удалить товар", callback_data=f"del_column_table")], [InlineKeyboardButton("\U00002795Добавить товар", callback_data=f"add_column_table")], [InlineKeyboardButton("\U0001F4CFИзменить количество", callback_data=f"count_table")], [InlineKeyboardButton("\U0001F519Назад", callback_data=f"exit_table")]])
    await app.send_message(chat_id=message.from_user.id, text="\U0000270FЧто вы хотите сделать?", reply_markup=markup)

async def add_column_table(client, message, name: str = True):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Добавление товара отменено", reply_markup=markup)
        await admin_table(message)
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    num = ""
    name_product = ""
    kolvo = ""
    text = message.text
    flag = True
    try:
        index = text.find("-")
        if index == -1:
            flag = False
        num = text[:(index-1)]
        text = text[(index+2):]
        index = text.find("-")
        if index == -1:
            flag = False
        name_product = text[:(index-1)]
        text = text[(index+2):]
        index = text.find("-")
        if index == -1:
            flag = False
        kolvo = text[:(index-1)]
        text = text[(index+2):]
        price = text
        if flag == False:
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"add_column_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
            return
        cursor.execute("SELECT * FROM Plants WHERE num = ?", (int(num),))
        is_num = cursor.fetchall()
        if is_num != []:
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"new_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.send_message(chat_id=message.from_user.id, text=f"\U0000274CТовар с таким номером уже существует в закупке **{is_num[0][1]}", reply_markup=markup)
            return
        cursor.execute('INSERT INTO Plants (num, purchage, name, count, price, book, ordered) VALUES (?, ?, ?, ?, ?, ?, ?)', (int(num), name, name_product, int(kolvo), int(price), 0, 0))
        connection.commit()
    except:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Создание таблицы отменено", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002714Повторить попытку", callback_data=f"add_column_table")], [InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка во введённых данных", reply_markup=markup)
        return
    cursor.execute("SELECT num, name, count, price FROM Plants WHERE purchage = ?", (name,))
    all_plants = cursor.fetchall()
    text = f"\U00002705Готово! Сейчас таблица для закупки **{name}** выглядит следующим образом:\n\n**Номер | Название растения | Количество | Цена**\n"
    for plant in all_plants:
        text += f"{plant[0]} | {plant[1]} | {plant[2]} | {plant[3]}р.\n"
    markup = ReplyKeyboardRemove()
    await app.send_message(chat_id=message.from_user.id, text="\U00002714Данные успешно прочитаны", reply_markup=markup)
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
    await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
    connection.close()

async def del_column_table(client, message):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Удаление товара отменено", reply_markup=markup)
        await admin_table(message)
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT num FROM Plants WHERE num = ?", (int(message.text),))
        need_num = cursor.fetchall()
        if (int(message.text),) in need_num:
            cursor.execute("SELECT purchage FROM Plants WHERE num = ?", (int(message.text),))
            purchage = cursor.fetchall()
            cursor.execute("SELECT * FROM orders WHERE product = ?", (int(message.text),))
            orders = cursor.fetchall()
            cursor.execute("SELECT price FROM Plants WHERE num = ?", (int(message.text),))
            price = cursor.fetchall()
            text_users = "\U0000203CУ некоторых пользователей из заказа был удалён оплаченный товар:\n\n"
            num = 1
            for order in orders:
                cursor.execute("SELECT price FROM Plants WHERE num = ?", (order[1],))
                price = cursor.fetchall()
                text_users += f"{num}. Пользователь {order[5]}, {order[3]} шт. на сумму {price[0][0] * order[3]}₽\n"
                cursor.execute("SELECT * FROM orders WHERE num = ?", (order[0],))
                is_order = cursor.fetchall()
                try:
                    text = f"\U0001F534\U00002702Из вашего заказа №{order[0]} был удалён товар **{order[2]}**, {order[3]} шт. на сумму {price[0][0] * order[3]}₽\n\n\U0001F4CCПричина: удаление товара из таблицы товаров продавцом"
                    if len(is_order) == 1:
                        text += "\n\n\U0000203CВ заказе товаров больше не осталось. Заказ полностью отменён"
                    await app.send_message(chat_id=order[4], text=text)
                except:
                    pass
                num += 1
            if text_users != "\U0000203CУ некоторых пользователей из заказа был удалён оплаченный товар:\n\n":
                await app.send_message(chat_id=message.from_user.id, text=text_users)
            cursor.execute("DELETE FROM orders WHERE product = ?", (int(message.text),))
            cursor.execute("DELETE FROM Plants WHERE num = ?", (int(message.text),))
            connection.commit()
            cursor.execute("SELECT num, name, count, price FROM Plants WHERE purchage = ?", (purchage[0][0],))
            all_plants = cursor.fetchall()
            text = f"\U00002705Готово! Теперь таблица для закупки **{purchage[0][0]}** выглядит следующим образом:\n\n**Номер | Название растения | Количество | Цена**\n"
            for plant in all_plants:
                text += f"{plant[0]} | {plant[1]} | {plant[2]} | {plant[3]}р.\n"
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U00002714Товар найден", reply_markup=markup)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
            await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
        else:
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U00002716Удаление товара отменено", reply_markup=markup)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"del_column_table")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет товара с таким номером", reply_markup=markup)
    except:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Удаление товара отменено", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"del_column_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет товара с таким номером", reply_markup=markup)
    connection.close()

async def count_table(client, message):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Изменения не сохранены", reply_markup=markup)
        await admin_table(message)
        return
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT num FROM Plants WHERE num = ?", (int(message.text),))
        need_num = cursor.fetchall()
        if (int(message.text),) in need_num:
            await app.send_message(chat_id=message.from_user.id, text=f"Введите новое количество для товара №**{need_num[0][0]}**")
            await pyrostep.register_next_step(message.from_user.id, count_table_do, kwargs={"num": int(message.text)})

        else:
            markup = ReplyKeyboardRemove()
            await app.send_message(chat_id=message.from_user.id, text="\U00002716Изменения не сохранены", reply_markup=markup)
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"count_table")]])
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет товара с таким номером", reply_markup=markup)
    except:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Изменения не сохранены", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"count_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CНет товара с таким номером", reply_markup=markup)
    connection.close()

async def count_table_do(client, message, num: int = True):
    if message.text == "\U0001F519Назад":
        await app.send_message(chat_id=message.from_user.id, text="\U0001F4CFВведите номер товара, количество которого нужно изменить")
        await app.delete_messages(chat_id=message.from_user.id, message_ids=message.id)
        await pyrostep.register_next_step(message.from_user.id, count_table)
    try:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("UPDATE Plants SET count = ? WHERE num = ?", (int(message.text), num))
        connection.commit()
        cursor.execute("SELECT purchage FROM Plants WHERE num = ?", (num,))
        purchage = cursor.fetchall()
        cursor.execute("SELECT num, name, count, price FROM Plants WHERE purchage = ?", (purchage[0][0],))
        all_plants = cursor.fetchall()
        text = f"\U00002705Готово! Теперь таблица для закупки **{purchage[0][0]}** выглядит следующим образом:\n\n**Номер | Название растения | Количество | Цена**\n"
        for plant in all_plants:
            text += f"{plant[0]} | {plant[1]} | {plant[2]} | {plant[3]}р.\n"
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Товар найден", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F3E0Главное меню", callback_data=f"exit_table")]])
        await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
        connection.close()
    except:
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Изменения не сохранены", reply_markup=markup)
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Повторить попытку", callback_data=f"count_table")]])
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CКоличество товара введено неверно", reply_markup=markup)

async def async_my_account(client, message):
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Изменить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**.\n\n\U00002139Данные для доставки:\n**ФИО**: {user[0][1]}\n**Адрес доставки**: {user[0][2]}\n**Телефон**: {user[0][3]}"
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Заполнить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**. Воспользуйтесь кнопками ниже\U00002B07"
        await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
        connection.close()

async def add_lk_name(client, message):
    if message.text == "\U0001F3E0Главное меню":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002716Изменения отменены", reply_markup=markup)
        await async_my_account("_", message)
        return
    elif message.text == "\U0001F519Назад":
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002702Введённые данные не были сохранены.\n\U0001F464Вы всегда можете заполнить информацию для отправки закзаов с помощью команды /account", reply_markup=markup)
        return
    else:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Temp_users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            cursor.execute("DELETE FROM Temp_users WHERE id = ?", (message.from_user.id,))
            connection.commit()
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        cursor.execute("INSERT INTO Temp_users (id, name) VALUES (?, ?)", (message.from_user.id, message.text))
        connection.commit()
        await app.send_message(chat_id=message.from_user.id, text="\U0000270FВведите удобный для получения заказов адрес", reply_markup=markup)
        connection.close()
        await pyrostep.register_next_step(message.from_user.id, add_lk_adress)

async def add_lk_adress(client, message):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F3E0Главное меню")]], resize_keyboard=True)
        await app.send_message(chat_id=message.from_user.id, text="\U0000270FВведите ваше ФИО", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, add_lk_name)
        return
    else:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F4CEОтправить мой номер", request_contact=True)], [KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        cursor.execute("UPDATE Temp_users SET adress = ? WHERE id = ?", (message.text, message.from_user.id))
        connection.commit()
        await app.send_message(chat_id=message.from_user.id, text="\U0000270FВведите номер вашего телефона или воспользуйтесь кнопкой\U00002B07", reply_markup=markup)
        connection.close()
        await pyrostep.register_next_step(message.from_user.id, add_lk_phone)

async def add_lk_phone(client, message):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    if message.contact != None:
        cursor.execute("UPDATE Temp_users SET telephone = ? WHERE id = ?", (f"+{message.contact.phone_number}", message.from_user.id))
        connection.commit()
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U00002705Сохранить")], [KeyboardButton(text="\U0001F519Назад")], [KeyboardButton(text="\U0000274CЗаполнить заново")]], resize_keyboard=True)
        cursor.execute("SELECT id, name, adress, telephone FROM Temp_users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        await app.send_message(chat_id=message.from_user.id, text=f"\U0001F50EПроверьте введённые данные:\n**ФИО:** {user[0][1]}\n**Адрес доставки:** {user[0][2]}\n**Телефон:** {user[0][3]}", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, save_lk_data)
        connection.close()
    else:
        if message.text == "\U0001F519Назад":
            markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
            await app.send_message(chat_id=message.from_user.id, text="\U0000270FВведите удобный для получения заказов адрес", reply_markup=markup)
            await pyrostep.register_next_step(message.from_user.id, add_lk_adress)
            connection.close()
            return
        else:
            cursor.execute("UPDATE Temp_users SET telephone = ? WHERE id = ?", (message.text, message.from_user.id))
            connection.commit()
            markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U00002705Сохранить")], [KeyboardButton(text="\U0001F519Назад")], [KeyboardButton(text="\U0000274CЗаполнить заново")]], resize_keyboard=True)
            cursor.execute("SELECT id, name, adress, telephone FROM Temp_users WHERE id = ?", (message.from_user.id,))
            user = cursor.fetchall()
            await app.send_message(chat_id=message.from_user.id, text=f"\U0001F50EПроверьте введённые данные:\n**ФИО:** {user[0][1]}\n**Адрес доставки:** {user[0][2]}\n**Телефон:** {user[0][3]}", reply_markup=markup)
            await pyrostep.register_next_step(message.from_user.id, save_lk_data)
            connection.close()

async def save_lk_data(client, message):
    if message.text == "\U0001F519Назад":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F4CEОтправить мой номер", request_contact=True)], [KeyboardButton(text="\U0001F519Назад")]], resize_keyboard=True)
        await app.send_message(chat_id=message.from_user.id, text="\U0000270FВведите номер вашего телефона или воспользуйтесь кнопкой\U00002B07", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, add_lk_phone)
        return
    elif message.text == "\U0000274CЗаполнить заново":
        markup = ReplyKeyboardMarkup([[KeyboardButton(text="\U0001F3E0Главное меню")]], resize_keyboard=True)
        await app.send_message(chat_id=message.from_user.id, text="\U0000270FВведите ваше ФИО", reply_markup=markup)
        await pyrostep.register_next_step(message.from_user.id, add_lk_name)
    else:
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT id, name, adress, telephone FROM Temp_users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (message.from_user.id,))
        user_check = cursor.fetchall()
        if user_check != []:
            cursor.execute("DELETE FROM Users WHERE id = ?", (message.from_user.id,))
            connection.commit()
        cursor.execute("INSERT INTO Users (id, name, adress, telephone) VALUES (?, ?, ?, ?)", (message.from_user.id, user[0][1], user[0][2], user[0][3]))
        connection.commit()
        markup = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002705Данные сохранены! Теперь вы сможете использовать их при оформлении заказа\U0001F389", reply_markup=markup)
        cursor.execute("SELECT id, name, adress, telephone FROM Users WHERE id = ?", (message.from_user.id,))
        user = cursor.fetchall()
        if user != []:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Изменить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**.\n\n\U00002139Данные для доставки:\n**ФИО**: {user[0][1]}\n**Адрес доставки**: {user[0][2]}\n**Телефон**: {user[0][3]}"
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F6D2Корзина", callback_data=f"basket")], [InlineKeyboardButton("\U0001F69BМои заказы", callback_data=f"my_orders")], [InlineKeyboardButton("\U00002699Заполнить данные для доставки", callback_data=f"change_my_data")]])
            text = f"\U0001F512Вы вошли в ваш **личный кабинет**. Воспользуйтесь кнопками ниже\U00002B07"
        await app.send_message(chat_id=message.from_user.id, text=text, reply_markup=markup)
        connection.close()

async def send_name(client, message, num: int = True):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    cursor.execute("UPDATE orders SET name = ? WHERE num = ?", (message.text, num))
    connection.commit()
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("Далее\U000027A1", callback_data=f"send_adress{num}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_name{num}")]])
    await app.send_message(chat_id=message.from_user.id, text=f"Ваше ФИО: **{message.text}**",reply_markup=markup)
    connection.close()

async def send_adress(client, message, num: int = True):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    cursor.execute("UPDATE orders SET adress = ? WHERE num = ?", (message.text, num))
    connection.commit()
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("Далее\U000027A1", callback_data=f"send_phone{num}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_adress{num}")]])
    await app.send_message(chat_id=message.from_user.id, text=f"Адрес доставки: **{message.text}**",reply_markup=markup)
    connection.close()

async def send_phone(client, message, num: int = True):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    if message.contact != None:
        cursor.execute("UPDATE orders SET telephone = ? WHERE num = ?", (f"+{message.contact.phone_number}", num))
        connection.commit()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Отправить данные", callback_data=f"send_data{num}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_phone{num}")]])
        markup2 = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Телефон проверен", reply_markup=markup2)
        await app.send_message(chat_id=message.from_user.id, text=f"Номер телефона: **+{message.contact.phone_number}**", reply_markup=markup)
    else:
        cursor.execute("UPDATE orders SET telephone = ? WHERE num = ?", (message.text, num))
        connection.commit()
        markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U00002705Отправить данные", callback_data=f"send_data{num}")], [InlineKeyboardButton("\U00002B05Назад", callback_data=f"send_phone{num}")]])
        markup2 = ReplyKeyboardRemove()
        await app.send_message(chat_id=message.from_user.id, text="\U00002714Телефон проверен", reply_markup=markup2)
        await app.send_message(chat_id=message.from_user.id, text=f"Номер телефона: **{message.text}**",reply_markup=markup)
    connection.close()

async def buy_done(client, message, num: int = True):
    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
    cursor = connection.cursor()
    cursor.execute("SELECT product, product_name, count, customer_id, customer_name, purchage FROM orders WHERE num = ?", (num,))
    orders = cursor.fetchall()
    if orders != []:
        if str(message.media) == "MessageMediaType.PHOTO":
            try:
                path = await app.download_media(message=message, file_name=f"C:/Users/Administrator/My_bots/Botanica_club_bot/{num}.jpg")
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F4CEПодтвердить оплату", callback_data=f"buy_confirm{num}")], [InlineKeyboardButton("\U0000274CОтклонить оплату", callback_data=f"buy_reject{num}")]])
                text = f"\U0001F4CEЧек к заказу №`{num}`\n\n\U0001F4E6Состав заказа:\n"
                spis_purchages = []
                for order in orders:
                    if order[5] not in spis_purchages:
                        spis_purchages.append(order[5])
                cost = 0
                spis_purchages = sorted(spis_purchages)
                for purchage in spis_purchages:
                    text += f"**Закупка {purchage}**\n"
                    number = 1
                    cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                    orderss = cursor.fetchall()
                    for orderr in orderss:
                        cursor.execute("SELECT price FROM Plants WHERE num = ?", (orderr[1],))
                        price = cursor.fetchall()
                        text += f"{number}. Товар {orderr[1]} **{orderr[2]}**. Кол-во: {orderr[3]} на сумму {price[0][0] * orderr[3]}₽.\n"
                        cost += price[0][0] * orderr[3]
                        number += 1
                text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
                await app.send_photo(chat_id=1333967466, photo=path, caption=text, reply_markup=markup)
                await app.send_photo(chat_id=993684230, photo=path, caption=text, reply_markup=markup)
                cursor.execute("UPDATE orders SET reciept = ?, status = ? WHERE num = ?", (path, 4, num))
                connection.commit()
                await app.send_message(chat_id=message.from_user.id, text=f"\U0001F7E2\U0001F4CEОжидайте подтверждения оплаты заказа №`{num}` модератором.")
            except:
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F4CEОтправить чек", callback_data=f"buy_done{num}")]]) 
                await app.send_message(chat_id=message.from_user.id, text="\U0000274CНеверный формат файла. Чек должен быть в формате скриншота или pdf-файла", reply_markup=markup)
        elif str(message.media) == "MessageMediaType.DOCUMENT" and message.document.file_name[-3:] == "pdf":
            try:
                path = await app.download_media(message=message, file_name=f"C:/Users/Administrator/My_bots/Botanica_club_bot/{num}.pdf")
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F4CEПодтвердить оплату", callback_data=f"buy_confirm{num}")], [InlineKeyboardButton("\U0000274CОтклонить оплату", callback_data=f"buy_reject{num}")]])
                text = f"\U0001F4CEЧек к заказу №`{num}`\n\n\U0001F4E6Состав заказа:\n\n"
                spis_purchages = []
                for order in orders:
                    if order[5] not in spis_purchages:
                        spis_purchages.append(order[5])
                cost = 0
                spis_purchages = sorted(spis_purchages)
                for purchage in spis_purchages:
                    text += f"**Закупка {purchage}**\n"
                    number = 1
                    cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay, status, purchage FROM orders WHERE num = ? and purchage = ?", (num, purchage))
                    orders = cursor.fetchall()
                    for order in orders:
                        cursor.execute("SELECT price FROM Plants WHERE num = ?", (order[1],))
                        price = cursor.fetchall()
                        text += f"{number}. Товар {order[1]} **{order[2]}**. Кол-во: {order[3]} на сумму {price[0][0] * order[3]}₽.\n"
                        cost += price[0][0] * order[3]
                        number += 1
                text += f"\n\U0001F4B5**Общая стоимость: {cost}₽**"
                await app.send_document(chat_id=1333967466, document=path, caption=text, reply_markup=markup)
                await app.send_document(chat_id=993684230, document=path, caption=text, reply_markup=markup)
                cursor.execute("UPDATE orders SET reciept = ?, status = ? WHERE num = ?", (path, 4, num))
                connection.commit()
                await app.send_message(chat_id=message.from_user.id, text="\U0001F7E2\U0001F4CEОжидайте подтверждения оплаты модератором.")
            except:
                markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F4CEОтправить чек", callback_data=f"buy_done{num}")]]) 
                await app.send_message(chat_id=message.from_user.id, text="\U0000274CНеверный формат файла. Чек должен быть в формате скриншота или pdf-файла", reply_markup=markup)
        else:
            markup = InlineKeyboardMarkup([[InlineKeyboardButton("\U0001F4CEОтправить чек", callback_data=f"buy_done{num}")]]) 
            await app.send_message(chat_id=message.from_user.id, text="\U0000274CНеверный формат файла. Чек должен быть в формате скриншота или pdf-файла", reply_markup=markup)
    else:
        await app.send_message(chat_id=message.from_user.id, text="\U0000274CОшибка. Пользователь отменил заказ или забронированный товар был удалён из списка товаров")
    connection.close()

async def view_table_orders(message, name_purchage):
        mess = await app.send_message(chat_id=message.from_user.id, text=f"\U0001F551Создаю таблицу для закупки **{name_purchage}**")
        connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
        cursor = connection.cursor()
        cursor.execute("SELECT num, name FROM Plants")
        plants = cursor.fetchall()
        head = ["Номер заказа", "Закупка", "Оплата", "Статус", "ID заказчика", "Контакт заказчика", "Доставка", "Тип доставки", "ФИО", "Адрес доставки", "Телефон"]
        plants_nums = []
        for plant in plants:
            head.append(plant[1])
            plants_nums.append(plant[0])
        cursor.execute("SELECT purchage FROM orders")
        pchg = cursor.fetchall()
        cursor.execute("SELECT num FROM orders WHERE purchage = ? ORDER BY num", (name_purchage,))
        nums = cursor.fetchall()
        spis_nums = []
        data = []
        spis_ids = []
        for num in nums:
            if num[0] not in spis_nums:
                spis_nums.append(num[0])
                cursor.execute("SELECT * FROM orders WHERE num = ? and purchage = ?", (num[0], name_purchage))
                orders = cursor.fetchall()
                if orders[0][5] not in spis_ids:
                    spis_ids.append(orders[0][5])
                    datapr = []
                    cursor.execute("SELECT * FROM orders WHERE num = ? ORDER BY purchage", (num[0],))
                    spis_num_need = cursor.fetchall()
                    need_num = 1
                    for spis_need_num_one in spis_num_need:
                        if spis_need_num_one[1] == name_purchage:
                            break
                        need_num += 1
                    flag = False
                    for spis_need_num_one in spis_num_need:
                        if spis_need_num_one[1] != name_purchage:
                            flag = True
                    if flag == True:
                        datapr.append(str(orders[0][0]) + "." + str(need_num))
                    else:
                        datapr.append(str(orders[0][0]))
                    datapr.append(name_purchage)
                    datapr.append(orders[0][7])
                    if orders[0][13] == 1:
                        datapr.append("В корзине")
                    elif orders[0][13] == 2:
                        datapr.append("Ожидает проверки наличия товара")
                    elif orders[0][13] == 3:
                        datapr.append("Ожидает оплаты")
                    elif orders[0][13] == 4:
                        datapr.append("Ожидает проверки оплаты")
                    elif orders[0][13] == 5:
                        datapr.append("Ожидает ввода данных для отправки")
                    elif orders[0][13] == 6:
                        datapr.append("Оплата подтверждена, заказ принят")
                    datapr.append(orders[0][5])
                    datapr.append(orders[0][6])
                    datapr.append(orders[0][8])
                    if orders[0][7] == "post":
                        if orders[0][13] == 1:
                            word_send = "Ускоренная"
                        else:
                            word_send = "Обычная"
                    else:
                        word_send = "-"
                    datapr.append(word_send)
                    datapr.append(orders[0][9])
                    datapr.append(orders[0][10])
                    datapr.append(orders[0][11])
                    product = {}
                    for order in orders:
                        product[f"{order[2]}"] = order[4]
                    for plant in plants_nums:
                        count = product.get(f"{plant}")
                        datapr.append(count)
                    cursor.execute("SELECT * FROM orders WHERE customer_id = ? and purchage = ?", (orders[0][5], name_purchage))
                    orders_id = cursor.fetchall()
                    data.append(datapr)
                    spis_nums_id = []
                    for order_id in orders_id:
                        if order_id[0] != num[0] and order_id[0] not in spis_nums_id:
                            spis_nums_id.append(order_id[0])
                            cursor.execute("SELECT * FROM orders WHERE num = ? and purchage = ?", (order_id[0], name_purchage))
                            orders_one_id = cursor.fetchall()
                            datapr = []
                            cursor.execute("SELECT * FROM orders WHERE num = ? ORDER BY purchage", (order_id[0],))
                            spis_num_need = cursor.fetchall()
                            need_num = 1
                            for spis_need_num_one in spis_num_need:
                                if spis_need_num_one[1] == name_purchage:
                                    break
                                need_num += 1
                            flag = False
                            for spis_need_num_one in spis_num_need:
                                if spis_need_num_one[1] != name_purchage:
                                    flag = True
                            if flag == True:
                                datapr.append(str(orders_one_id[0][0]) + "." + str(need_num))
                            else:
                                datapr.append(str(orders_one_id[0][0]))
                            datapr.append(name_purchage)
                            datapr.append(orders_one_id[0][7])
                            if orders_one_id[0][13] == 1:
                                datapr.append("В корзине")
                            elif orders_one_id[0][13] == 2:
                                datapr.append("Ожидает проверки наличия товара")
                            elif orders_one_id[0][13] == 3:
                                datapr.append("Ожидает оплаты")
                            elif orders_one_id[0][13] == 4:
                                datapr.append("Ожидает проверки оплаты")
                            elif orders_one_id[0][13] == 5:
                                datapr.append("Ожидает ввода данных для отправки")
                            elif orders_one_id[0][13] == 6:
                                datapr.append("Оплата подтверждена, заказ принят")
                            datapr.append(orders_one_id[0][5])
                            datapr.append(orders_one_id[0][6])
                            datapr.append(orders_one_id[0][8])
                            if orders_one_id[0][7] == "post":
                                if orders_one_id[0][13] == 1:
                                    word_send = "Ускоренная"
                                else:
                                    word_send = "Обычная"
                            else:
                                word_send = "-"
                            datapr.append(word_send)
                            datapr.append(orders_one_id[0][9])
                            datapr.append(orders_one_id[0][10])
                            datapr.append(orders_one_id[0][11])
                            product = {}
                            for order_one in orders_one_id:
                                product[f"{order_one[2]}"] = order_one[4]
                            for plant in plants_nums:
                                count = product.get(f"{plant}")
                                datapr.append(count)
                            data.append(datapr)
        wb = Workbook()
        ws = wb.active
        ws.append(head)
        for i in range(len(data)):
            ws.append(data[i])
        for i in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            ws[f'{i}1'].font = Font(bold=True)
        ws[f'AA1'].font = Font(bold=True)
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 50
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 20
        ws.column_dimensions['P'].width = 20
        ws.column_dimensions['Q'].width = 20
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 20
        ws.column_dimensions['T'].width = 20
        ws.column_dimensions['U'].width = 20
        ws.column_dimensions['V'].width = 20
        ws.column_dimensions['W'].width = 20
        ws.column_dimensions['X'].width = 20
        ws.column_dimensions['Y'].width = 20
        ws.column_dimensions['Z'].width = 20
        ws.column_dimensions['AA'].width = 20
        ws.column_dimensions['AB'].width = 20
        ws.column_dimensions['AC'].width = 20
        ws.column_dimensions['AD'].width = 20
        ws.column_dimensions['AE'].width = 20
        ws.column_dimensions['AF'].width = 20
        ws.column_dimensions['AG'].width = 20
        for col in ws.columns:
            for cell in col:
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj
        connection.close()
        wb.save(f"C:/Users/Administrator/My_bots/Botanica_club_bot/{name_purchage}.xlsx")
        await app.send_document(chat_id=message.from_user.id, document=f"C:/Users/Administrator/My_bots/Botanica_club_bot/{name_purchage}.xlsx", caption=f"\U0001F4CEУдалённая таблица заказов для закупки {name_purchage}")
        await app.delete_messages(chat_id=message.from_user.id, message_ids=mess.id)

@app.on_message(filters.chat(chats=[-1001789882072]))
def posts_orders(client, message):
    if message.reply_to_message != None and message.from_user != None and message.text != None:
        if message.reply_to_message.text != None:
            mess_text = message.reply_to_message.text
        elif message.reply_to_message.caption != None:
            mess_text = message.reply_to_message.caption
        else:
            mess_text = ""
        if "#" in mess_text and "Артикул №" in mess_text:
            if (message.text).isdigit():
                connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
                cursor = connection.cursor()
                index = (mess_text).find("Артикул №")
                ind = ""
                for s in (mess_text)[index + 9:]:
                    if s.isdigit():
                        ind += s
                    else:
                        break
                ind = int(ind)
                index_purchage = (mess_text).find("#")
                text_find = (mess_text)[index_purchage:]
                index_purchage2 = text_find.find("\n")
                purchage = (mess_text)[index_purchage:index_purchage2 + index_purchage]
                cursor.execute("SELECT num, name, count, book FROM Plants WHERE num = ? and purchage = ?", (ind, purchage))
                plant = cursor.fetchall()
                if plant != []:
                    if (message.text).isdigit() and plant[0][2] - plant[0][3] >= int(message.text):
                        cursor.execute('SELECT num FROM orders WHERE customer_id = ? and status = ?', (message.from_user.id, 1))
                        cust_id = cursor.fetchall()
                        if cust_id == []:
                            cursor.execute('SELECT num FROM orders')
                            cust_id = cursor.fetchall()
                            if cust_id == []:
                                cust_id = 1
                            else:
                                max_num = 0
                                for id in cust_id:
                                    if id[0] > max_num:
                                        max_num = id[0]
                                cust_id = max_num
                                cust_id += 1
                        else:
                            cust_id = cust_id[0][0]
                        cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (plant[0][3] + int(message.text), ind))
                        cursor.execute('SELECT count FROM orders WHERE customer_id = ? and status = ? and product = ?', (message.from_user.id, 1, ind))
                        count_same = cursor.fetchall()
                        if count_same == []:
                            if message.from_user.username != None:
                                cursor.execute('INSERT INTO orders (num, product, product_name, count, customer_id, customer_name, status, purchage) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (cust_id, plant[0][0], plant[0][1], int(message.text), message.from_user.id, "@" + message.from_user.username, 1, purchage))
                            else:
                                cursor.execute('INSERT INTO orders (num, product, product_name, count, customer_id, customer_name, status, purchage) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', (cust_id, plant[0][0], plant[0][1], int(message.text), message.from_user.id, f"<a href='tg://user?id={message.from_user.id}'>{message.from_user.first_name}</a>", 1, purchage))
                        else:
                            cursor.execute('UPDATE orders SET count = ? WHERE customer_id = ? and status = ? and product = ?', (count_same[0][0] + int(message.text), message.from_user.id, 1, ind))
                        connection.commit()
                        cursor.execute("SELECT num FROM orders WHERE num = ? and product = ? and product_name = ? and count = ? and customer_id = ?", (cust_id, plant[0][0], plant[0][1], int(message.text), message.from_user.id))
                        num = cursor.fetchall()
                        try:
                            app.send_message(chat_id=message.from_user.id, text=f'\U00002705\U0001F6D2Вы успешно добавили в корзину {message.text} шт. товара №{plant[0][0]}: **{plant[0][1]}**.\n\n\U0001F6D2Посмотреть корзину и сделать заказ можно с помощью команды /basket')
                        except:
                            app.send_message(chat_id=message.chat.id, text="\U0000203CЧтобы добавить товар в корзину, напишите боту @Botanica_club_bot и попробуйте снова", reply_to_message_id=message.id)
                            cursor.execute("DELETE FROM orders WHERE num = ?", (num[-1][0],))
                            connection.commit()
                            cursor.execute("SELECT book FROM Plants WHERE num = ?", (plant[0][0],))
                            count = cursor.fetchall()
                            cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (plant[0][3], plant[0][0]))
                            connection.commit()
                    elif (message.text).isdigit() and plant[0][2] - plant[0][3] < int(message.text):
                        try:
                            app.send_message(chat_id=message.from_user.id, text=f"\U0000274CВы попытались добавить в корзину {message.text} шт. товара №{plant[0][0]}: **{plant[0][1]}**.\n\n\U00002705В наличии {plant[0][2] - plant[0][3]} шт.")
                        except:
                            app.send_message(chat_id=message.chat.id, text="\U0000203CЧтобы добавить товар в корзину, напишите боту @Botanica_club_bot и попробуйте снова", reply_to_message_id=message.id)
                connection.close()
            elif message.text == "Отказ" or message.text == "отказ":
                try:
                    connection = sqlite3.connect("C:/Users/Administrator/My_bots/Botanica_club_bot/Plants.db")
                    cursor = connection.cursor()
                    index = (mess_text).find("Артикул №")
                    ind = ""
                    for s in (mess_text)[index + 9:]:
                        if s.isdigit():
                            ind += s
                        else:
                            break
                    ind = int(ind)
                    index_purchage = (mess_text).find("#")
                    text_find = (mess_text)[index_purchage:]
                    index_purchage2 = text_find.find("\n")
                    purchage = (mess_text)[index_purchage:index_purchage2 + index_purchage]
                    cursor.execute("SELECT num, name, count, book, price FROM Plants WHERE num = ?", (ind,))
                    plant = cursor.fetchall()
                    if plant != []:
                        cursor.execute("SELECT num, product, product_name, count, customer_id, customer_name, is_pay FROM orders WHERE customer_id = ? and product = ? and purchage = ?", (message.from_user.id, ind, purchage))
                        orders = cursor.fetchall()
                        if orders == []:
                            try:
                                app.send_message(chat_id=message.from_user.id, text=f"\U0000274CОшибка отказа от товара. В ваших заказах нет такого товара")
                            except:
                                app.send_message(chat_id=message.chat.id, text="\U0000203CДля отказа от товара напишите боту @Botanica_club_bot и попробуйте снова", reply_to_message_id=message.id)
                            return
                        cursor.execute("SELECT book FROM Plants WHERE num = ?", (orders[0][1],))
                        count = cursor.fetchall()
                        for order in orders:
                            if order[6] != 1:
                                cursor.execute("DELETE FROM orders WHERE num = ? and product = ?", (order[0], ind))
                                connection.commit()
                                cursor.execute("UPDATE Plants SET book = ? WHERE num = ?", (count[0][0] - order[3], order[1]))
                                connection.commit()
                                try:
                                    app.send_message(chat_id=message.from_user.id, text=f"\U0001F534\U0001F6D2В заказе №`{order[0]}` товар **{order[2]}**, {order[3]} шт. на сумму {plant[0][4] * order[3]}₽ успешно удалён")
                                except:
                                    app.send_message(chat_id=message.chat.id, text="\U0000203CДля отказа от товара напишите боту @Botanica_club_bot", reply_to_message_id=message.id)
                except:
                    try:
                        app.send_message(chat_id=message.from_user.id, text=f"\U0000274CОшибка отказа от товара. В ваших заказах нет такого товара")
                    except:
                        app.send_message(chat_id=message.chat.id, text="\U0000203CЧтобы добавить товар в корзну, напишите боту @Botanica_club_bot и попробуйте снова", reply_to_message_id=message.id)
                connection.close()

app.run()   