import telebot
bot = telebot.TeleBot('BOT TOKEN')
from telebot import types
import sqlite3
from datetime import timedelta, date
import pandas as pd
import schedule
import threading
import time as tm
from datetime import time
from openpyxl import Workbook
from openpyxl.styles import Font

@bot.message_handler(commands=['start'])
def startbot(message):
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
    cursor = connection.cursor()
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS Users (
    id INTEGER PRIMARY KEY,
    nick TEXT,
    phone TEXT,
    name TEXT
    )
    ''')
    cursor.execute('SELECT * FROM Users WHERE id = ?', (message.chat.id,))
    info = cursor.fetchall()
    if info == []:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button = types.KeyboardButton(text="Отправить номер телефона\U0001F4DE", request_contact=True)
        markup.add(button)
        mess = "Бот Mafia VIP приветствует вас\U0001F44B\n\nЯ помогу забронировать столик для игры в мафию, а также найти для этого компанию!\U0001F4C5🔎\n\nДля продолжения отправьте, пожалуйста, ваш номер телефона с помощью кнопки\U00002B07"
        cursor.execute('INSERT INTO Users (id, nick) VALUES (?, ?)', (message.chat.id, ("@" + message.chat.username)))
        connection.commit()
        msg = bot.send_message(message.chat.id, mess, reply_markup=markup)
        bot.register_next_step_handler(msg, contact_add)
    elif info[0][3] == None:
        cursor.execute('DELETE FROM Users WHERE id = ?', (message.chat.id,))
        connection.commit()
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button = types.KeyboardButton(text="Отправить номер телефона\U0001F4DE", request_contact=True)
        markup.add(button)
        mess = "Бот Mafia VIP приветствует вас\U0001F44B\n\nЯ помогу забронировать столик для игры в мафию, а также найти для этого компанию!\U0001F4C5🔎\n\nДля продолжения отправьте, пожалуйста, ваш номер телефона с помощью кнопки\U00002B07"
        cursor.execute('INSERT INTO Users (id, nick) VALUES (?, ?)', (message.chat.id, ("@" + message.chat.username)))
        connection.commit()
        msg = bot.send_message(message.chat.id, mess, reply_markup=markup)
        bot.register_next_step_handler(msg, contact_add)
    else:
        markup = types.InlineKeyboardMarkup()
        button_book = types.InlineKeyboardButton(text="Забронировать стол\U0001F195\U00002705", callback_data="book")
        button_company = types.InlineKeyboardButton(text="Найти компанию🔎", callback_data="find_company")
        markup.add(button_book)
        markup.add(button_company)
        mess = "Бот Mafia VIP приветствует вас!\U0001F44B\n\nДля бронирования столика воспользуйтесь кнопками\U0001F4CD\n\nДля связи с оператором введите\n/help\U00002B07"
        bot.send_message(message.chat.id, mess, reply_markup=markup) 
    connection.close()
    if message.chat.id == 1333967466:
        markup = types.InlineKeyboardMarkup()
        button1 = types.InlineKeyboardButton(text="Рассылка", callback_data="sending")
        button = types.InlineKeyboardButton(text="Посмотреть записи", callback_data="view")
        button2 = types.InlineKeyboardButton(text="Все пользователи", callback_data="view_all")
        button3 = types.InlineKeyboardButton(text="Добавить бронь", callback_data="add_admin")
        button4 = types.InlineKeyboardButton(text="Удалить бронь", callback_data="del_admin")
        button5 = types.InlineKeyboardButton(text="Посмотреть записи за месяц", callback_data="see_month")
        markup.add(button1)
        markup.add(button, button2)
        markup.add(button3, button4)
        markup.add(button5)
        bot.send_message(message.chat.id, "Меню администратора\U00002B07", reply_markup=markup)

@bot.message_handler(commands=['help'])
def helpbot(message):
    mess = f"По всем вопросам и проблемам используйте ссылку ниже:👨‍💻\U0001F447\n<a href='https://t.me/nekitkl'>Написать оператору</a>"
    bot.send_message(message.chat.id, mess, parse_mode="html")


def contact_add(message):
    if message.content_type == "contact":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Users SET phone = ? where id = ?', (message.contact.phone_number, message.chat.id))
        connection.commit()
        connection.close()
        msg = bot.send_message(message.chat.id, "Спасибо!\U00002705 Представьтесь, пожалуйста", reply_markup=types.ReplyKeyboardRemove())
        bot.register_next_step_handler(msg, name_add)
    else:
        msg = bot.send_message(message.chat.id, "Ошибка!\U0000274C Пожалуйста, воспользуйтесь кнопкой ниже\U00002B07")
        bot.register_next_step_handler(msg, contact_add)

def name_add(message):
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Users SET name = ? where id = ?', (message.text, message.chat.id))
        connection.commit()
        connection.close()
        msg = bot.send_message(message.chat.id, "Вы зарегистрированы. Спасибо!\U00002705")
        startbot(message)

@bot.callback_query_handler(func=lambda call:True)
def response(function_call):
    if function_call.data == "book":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[:7]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_back = types.InlineKeyboardButton(text="Назад\U00002B05\U00002B05", callback_data="main")
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_1")
        markup.add(button_back, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_1":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[7:14]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_2")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="book")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_2":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[14:21]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_3")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="next_1")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_3":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[21:28]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_4")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="next_2")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_4":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[28:35]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_5")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="next_3")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_5":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[35:42]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_6")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="next_4")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_6":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[42:49]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="next_7")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="next_5")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "next_7":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[49:]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_when{day[0]}")
            markup.add(button_day)
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="next_6")
        markup.add(button_prev)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "day_when" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT num, time, client1 from day_db{str((function_call.data)[8:])}')
        data = cursor.fetchall()
        cursor.execute('SELECT day FROM Shedule WHERE num = ?', (str(function_call.data)[8:],))
        day = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect(f'C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS Writes (
        id INTEGER PRIMARY KEY,
        date TEXT,
        time TEXT,
        hours INTEGER,
        name TEXT
        )
        ''')
        try:
            cursor.execute('SELECT * FROM Writes WHERE id = ?', (int(function_call.message.chat.id),))
        except:
            pass
        ifwrite = cursor.fetchall()
        if ifwrite != []:
            cursor.execute('DELETE FROM Writes WHERE id = ?', (function_call.message.chat.id,))
        connection.commit()
        cursor.execute('INSERT INTO Writes (id, date) VALUES (?, ?)', (function_call.message.chat.id, day[0][0]))
        connection.commit()
        markup = types.InlineKeyboardMarkup()
        spis_buttons = []
        for dat in range(len(data)):
            if data[dat][2] == None:
                count = 0
                if dat + 5 < len(data):
                    for i in range(1, 5):
                        if data[dat + i][2] == None:
                            count += 1
                        else:
                            break
                else:
                    fl = True
                    for i in range(1, (len(data)-dat)):
                        if data[dat + i][2] == None:
                            count += 1
                        else:
                            fl = False
                            break
                    if fl == True:
                        count = 4
                if count != 0:
                    spis_buttons.append(types.InlineKeyboardButton(text=f"{data[dat][1][:5]}, лимит: {str(count)} час(а)", callback_data=f"time{data[dat][1][:5]}{str(count)}"))
        for i in range(0, len(spis_buttons)-1, 2):
            markup.add(spis_buttons[i], spis_buttons[i+1])
        connection.close()    
        mess = "В какое время вы нас посетите?\U0001F551"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
    elif "time" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Writes SET time = ? WHERE id = ?', ((function_call.data)[4:9], function_call.message.chat.id))
        connection.commit()
        markup = types.InlineKeyboardMarkup()
        for i in range(1, int((function_call.data)[9:]) + 1):
            button = types.InlineKeyboardButton(text=f"{i} час(а)", callback_data=f"tme_lst{i}")
            markup.add(button)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text="Сколько часов бронируем?\U0001F550", reply_markup=markup, message_id=function_call.message.id)
    elif "tme_lst" in function_call.data:
        connection = sqlite3.connect(f'C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Writes SET hours = ? WHERE id = ?', ((function_call.data)[7:], function_call.message.chat.id))
        connection.commit()
        cursor.execute('SELECT * FROM Writes WHERE id = ?', (function_call.message.chat.id,))
        data = cursor.fetchall()
        mess = f"Проверьте правильность записи:\U0001F50E\n\nДата: {data[0][1]}\nВремя: {data[0][2][:5]}\nКоличество часов: {data[0][3]}"
        connection.close()
        markup = types.InlineKeyboardMarkup()
        button_yes = types.InlineKeyboardButton(text="Всё верно\U00002705", callback_data="payment")
        button_back = types.InlineKeyboardButton(text="Ошибка\U0000274C", callback_data="main")
        markup.add(button_yes)
        markup.add(button_back)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "payment":
        markup = types.InlineKeyboardMarkup()
        button_yes = types.InlineKeyboardButton(text="Оплачено\U00002705", callback_data="done_pay")
        markup.add(button_yes)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text="Функция в разработке", reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "main":
        markup = types.InlineKeyboardMarkup()
        button_book = types.InlineKeyboardButton(text="Забронировать стол\U0001F195\U00002705", callback_data="book")
        button_company = types.InlineKeyboardButton(text="Найти компанию🔎", callback_data="find_company")
        markup.add(button_book)
        markup.add(button_company)
        mess = "Бот Mafia VIP приветствует вас!\U0001F44B\n\nДля бронирования столика воспользуйтесь кнопками\U0001F4CD\n\nДля связи с оператором введите\n/help\U00002B07\n\nДля настройки бота используйте /settings\U00002699"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "admin_main":
        markup = types.InlineKeyboardMarkup()
        button1 = types.InlineKeyboardButton(text="Рассылка", callback_data="sending")
        button = types.InlineKeyboardButton(text="Посмотреть записи", callback_data="view")
        button2 = types.InlineKeyboardButton(text="Все пользователи", callback_data="view_all")
        button3 = types.InlineKeyboardButton(text="Добавить бронь", callback_data="add_admin")
        button4 = types.InlineKeyboardButton(text="Удалить бронь", callback_data="del_admin")
        markup.add(button1)
        markup.add(button, button2)
        markup.add(button3, button4)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text="Меню администратора\U00002B07", reply_markup=markup, message_id=function_call.message.id)
    elif function_call.data == "done_pay":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * FROM Writes WHERE id = ?', (function_call.message.chat.id,))
        reg = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT num from Shedule WHERE day = ?', (reg[0][1],))
        num = cursor.fetchall()
        num = num[0][0]
        cursor.execute(f'SELECT num from day_db{num} WHERE time = ?', ((reg[0][2] + ":00"),))
        num_time = cursor.fetchall()
        num_time = num_time[0][0]
        cursor.execute(f'UPDATE day_db{num} SET pay = ? WHERE num = ?', (True, (num_time)))
        cursor.execute(f'UPDATE day_db{num} SET administrator = ? WHERE num = ?', (False, (num_time)))
        for i in range(reg[0][3]):
            cursor.execute(f'UPDATE day_db{num} SET pay = ? WHERE num = ?', (True, (num_time + i)))
            cursor.execute(f'UPDATE day_db{num} SET administrator = ? WHERE num = ?', (False, (num_time + i)))
            cursor.execute(f'UPDATE day_db{num} SET client1 = ? WHERE num = ?', (reg[0][0], (num_time + i)))
        connection.commit()
        num_time += (i+1)
        for i in range(num_time, 15):
            cursor.execute(f'SELECT time, client1 from day_db{num} WHERE num = ?', (i,))
            current_reg = cursor.fetchall()
            if current_reg[0][1] == None:
                c_time = current_reg[0][0]
                hr = int(c_time[:2])
                mnte = int(c_time[3:5])
                if mnte != 40:
                    mnte += 20
                else:
                    mnte = 0
                    hr += 1
                if mnte == 0:
                    mnte = "00"
                times = str(hr) + ":" + str(mnte) + ":" + "00"
                cursor.execute(f'UPDATE day_db{num} SET time = ? WHERE num = ?', (times, i))
                connection.commit()
            else:
               break
        cursor.execute('SELECT day FROM Shedule WHERE num = ?', (num,))
        dating = cursor.fetchall()
        mess = f"Ждём вас {reg[0][1]} в {reg[0][2]} на {reg[0][3]} час(а)\U0001F917"
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
        cursor = connection.cursor()
        cursor.execute('SELECT nick, phone, name FROM Users WHERE id = ?', (function_call.message.chat.id,))
        user_info = cursor.fetchall()
        bot.send_message(chat_id=1333967466, text=f"Новая запись!\nДата: {dating[0][0]}\nВремя: {reg[0][2]}\nНик в ТГ: {user_info[0][0]}\nТелефон: {user_info[0][1]}\nИмя: {user_info[0][2]}")
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, message_id=function_call.message.id)
        markup = types.InlineKeyboardMarkup()
        button_book = types.InlineKeyboardButton(text="Забронировать стол\U0001F195\U00002705", callback_data="book")
        button_company = types.InlineKeyboardButton(text="Найти компанию🔎", callback_data="find_company")
        markup.add(button_book)
        markup.add(button_company)
        mess = "Для бронирования столика воспользуйтесь кнопками\U0001F4CD\n\nДля связи с оператором введите\n/help\U00002B07"
        bot.send_message(function_call.message.chat.id, mess, reply_markup=markup) 
        connection.close()
    elif function_call.data == "find_company":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[:7]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_back = types.InlineKeyboardButton(text="Назад\U00002B05\U00002B05", callback_data="main")
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_1")
        markup.add(button_back, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_1":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[7:14]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_2")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="find_company")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_2":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[14:21]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_3")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="nexxt_1")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_3":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[21:28]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_4")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="nexxt_2")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_4":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[28:35]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_5")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="nexxt_3")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_5":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[35:42]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_6")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="nexxt_4")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_6":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[42:49]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="nexxt_7")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="nexxt_5")
        markup.add(button_prev, button_next)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "nexxt_7":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[49:]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"when_fc{day[0]}")
            markup.add(button_day)
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="nexxt_6")
        markup.add(button_prev)
        mess = "Какого числа вы хотите прийти к нам?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "when_fc" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT num, time, client1 from day_db{str((function_call.data)[7:])}')
        data = cursor.fetchall()
        cursor.execute(f'SELECT day from Shedule WHERE num = ?', (str((function_call.data)[7:]),))
        date = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Find_company.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 from find_db{str((function_call.data)[7:])}')
        users = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        fl = False
        count = 0
        for i in range(1, 4):
            if data[i][2] == None:
                fl = True
        for j in range(len(users[0])):
            if users[0][j] != None:
                count += 1
        if fl == True:
            button_day = types.InlineKeyboardButton(text=f"Утро, ищут компанию: {count}", callback_data=f"find_final_add_m{(function_call.data)[7:]}")
            markup.add(button_day)
        else:
            spis = []
            for k in range(len(users[0])):
                if users[0][k] != None:
                    if users[0][k] not in spis:
                        spis.append(users[0][k])
                        try:
                            bot.send_message(chat_id=int(users[0][k]), text=f"Бот приветствует вас!\U0001F44B Вы оставляли заявку на поиск компании на {date[0][0]} К сожалению, в указанный вами промежуток времени записей не осталось\U0000274C Пожалуйста, выберите другое время\U00002705\U0001F553")
                            cursor.execute(f'UPDATE find_db{str((function_call.data)[7:])} SET client{str(k+1)} = NULL WHERE num = ?', (1,))
                            connection.commit()
                            cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 from find_db{str((function_call.data)[7:])}')
                            users = cursor.fetchall()
                        except:
                            pass
                    else:
                        cursor.execute(f'UPDATE find_db{str((function_call.data)[7:])} SET client{str(k+1)} = NULL WHERE num = ?', (1,))
                        connection.commit()
        fl = False
        count = 0
        for i in range(4, 9):
            if data[i][2] == None:
                fl = True
        for j in range(len(users[1])):
            if users[1][j] != None:
                count += 1
        if fl == True:
            button_day = types.InlineKeyboardButton(text=f"День, ищут компанию: {count}", callback_data=f"find_final_add_d{(function_call.data)[7:]}")
            markup.add(button_day)
        else:
            spis = []
            for k in range(len(users[1])):
                if users[1][k] != None:
                    if users[1][k] not in spis:
                        spis.append(users[1][k])
                        try:
                            bot.send_message(chat_id=int(users[1][k]), text=f"Бот приветствует вас!\U0001F44B Вы оставляли заявку на поиск компании на {date[0][0]} К сожалению, в указанный вами промежуток времени записей не осталось\U0000274C Пожалуйста, выберите другое время\U00002705\U0001F553")
                            cursor.execute(f'UPDATE find_db{str((function_call.data)[7:])} SET client{str(k+1)} = NULL WHERE num = ?', (2,))
                            connection.commit()
                            cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 from find_db{str((function_call.data)[7:])}')
                            users = cursor.fetchall()
                        except:
                            pass
                    else:
                        cursor.execute(f'UPDATE find_db{str((function_call.data)[7:])} SET client{str(k+1)} = NULL WHERE num = ?', (2,))
                        connection.commit()
        fl = False
        count = 0
        for i in range(9, 14):
            if data[i][2] == None:
                fl = True
        for j in range(len(users[2])):
            if users[2][j] != None:
                count += 1
        if fl == True:
            button_day = types.InlineKeyboardButton(text=f"Вечер, ищут компанию: {count}", callback_data=f"find_final_add_e{(function_call.data)[7:]}")
            markup.add(button_day)
        else:
            spis = []
            for k in range(len(users[2])):
                if users[2][k] != None:
                    if users[2][k] not in spis:
                        spis.append(users[2][k])
                        try:
                            bot.send_message(chat_id=int(users[2][k]), text=f"Бот приветствует вас!\U0001F44B Вы оставляли заявку на поиск компании на {date[0][0]} К сожалению, в указанный вами промежуток времени записей не осталось\U0000274C Пожалуйста, выберите другое время\U00002705\U0001F553")
                            cursor.execute(f'UPDATE find_db{str((function_call.data)[7:])} SET client{str(k+1)} = NULL WHERE num = ?', (3,))
                            connection.commit()
                        except:
                            pass
                    else:
                        cursor.execute(f'UPDATE find_db{str((function_call.data)[7:])} SET client{str(k+1)} = NULL WHERE num = ?', (3,))
                        connection.commit()
        mess = "Когда вам будет удобно нас посетить?\U0001F551"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "find_final_add" in function_call.data:
        markup = types.InlineKeyboardMarkup(row_width=3)
        for i in range(1, 6, 3):
            button_1 = types.InlineKeyboardButton(text=f"{i}", callback_data=f"clients_kolvo{i}{(function_call.data)[15:]}")
            button_2 = types.InlineKeyboardButton(text=f"{i+1}", callback_data=f"clients_kolvo{i+1}{(function_call.data)[15:]}")
            button_3 = types.InlineKeyboardButton(text=f"{i+2}", callback_data=f"clients_kolvo{i+2}{(function_call.data)[15:]}")
            markup.add(button_1, button_2, button_3)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text="Сколько вас человек?", reply_markup=markup, message_id=function_call.message.id)
    elif "clients_kolvo" in function_call.data:
        connection2 = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor2 = connection2.cursor()
        cursor2.execute('SELECT day FROM Shedule WHERE num = ?', (int((function_call.data)[15:]),))
        dt = cursor2.fetchall()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Find_company.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 from find_db{str((function_call.data)[15:])}')
        users = cursor.fetchall()
        if (function_call.data)[14] == "m":
            numer = 1
        elif (function_call.data)[14] == "d":
            numer = 2
        elif (function_call.data)[14] == "e":
            numer = 3
        for k in range(int((function_call.data)[13])):
            for i in range(len(users[0])):
                if users[numer-1][i] == None:
                    cursor.execute(f'UPDATE find_db{(function_call.data)[15:]} SET client{i+1} = ? WHERE num = ?', (function_call.message.chat.id, numer))
                    connection.commit()
                    cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 from find_db{str((function_call.data)[15:])}')
                    users = cursor.fetchall()
                    break
        cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12, time from find_db{str((function_call.data)[15:])}')
        clients = cursor.fetchall()
        if clients[numer-1][11] != None:
            spis_clients = []
            cursor2.execute(f'SELECT time, client1 FROM day_db{(function_call.data)[15:]}')
            timmes = cursor2.fetchall()
            connection2.close()
            if clients[numer-1][12] == "Утро":
                for i in range(3):
                    if timmes[i][1] == None:
                        tm = timmes[i][0]
                        break
            elif clients[numer-1][12] == "День":
                for i in range(3, 8):
                    if timmes[i][1] == None:
                        tm = timmes[i][0]
                        break
            elif clients[numer-1][12] == "Вечер":
                for i in range(8, 14):
                    if timmes[i][1] == None:
                        tm = timmes[i][0]
                        break
            for i in range(len(clients[numer-1]) - 1):
                if clients[numer-1][i] not in spis_clients:
                    spis_clients.append(clients[numer-1][i])
                    markup = types.InlineKeyboardMarkup()
                    button_reg = types.InlineKeyboardButton(text="Я приду!\U00002705", callback_data=f"end_find_reg{tm[:5]}{(function_call.data)[15:]}")
                    markup.add(button_reg)
                    mess = f"Компания нашлась!\U00002705\n\nВы оставляли заявку на поиск компании\U0001F4DD\nДата: {dt[0][0]} Время: {tm[:5]}\n\nНажмите на кнопку ниже, если посетите нас\U00002B07"
                    try:
                        bot.send_message(chat_id=int(clients[numer-1][i]), text=mess, reply_markup=markup)
                    except:
                        pass
                    cursor.execute(f'UPDATE find_db{(function_call.data)[15:]} SET client{str(i+1)} = NULL WHERE time = ?', (clients[numer-1][12],))
                    connection.commit()
                else:
                    cursor.execute(f'UPDATE find_db{(function_call.data)[15:]} SET client{str(i+1)} = NULL WHERE time = ?', (clients[numer-1][12],))
                    connection.commit()   
            connection2.close()
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=f"Запись на {dt[0][0]} создана!\U00002705 Ждите уведомлений\U0001F514", message_id=function_call.message.id)
    elif "end_find_reg" in function_call.data:
        mess = "Сколько вас человек?"
        markup = types.InlineKeyboardMarkup()
        button_1 = types.InlineKeyboardButton(text="1", callback_data=f"wr_db_find1{(function_call.data)[12:]}")
        button_2 = types.InlineKeyboardButton(text="2", callback_data=f"wr_db_find2{(function_call.data)[12:]}")
        button_3 = types.InlineKeyboardButton(text="3", callback_data=f"wr_db_find3{(function_call.data)[12:]}")
        markup.add(button_1, button_2, button_3)
        button_4 = types.InlineKeyboardButton(text="4", callback_data=f"wr_db_find4{(function_call.data)[12:]}")
        button_5 = types.InlineKeyboardButton(text="5", callback_data=f"wr_db_find5{(function_call.data)[12:]}")
        button_6 = types.InlineKeyboardButton(text="6", callback_data=f"wr_db_find6{(function_call.data)[12:]}")
        markup.add(button_4, button_5, button_6)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, message_id=function_call.message.id, reply_markup=markup)
    elif "wr_db_find" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 FROM day_db{(function_call.data)[16:]} WHERE time = ?', (((function_call.data)[11:16] + ":00"),))
        clients = cursor.fetchall()
        flag = True
        for i in range(int((function_call.data)[10])):
            fl = True
            k = 0
            for cl in clients[0]:
                if cl == None:
                    cursor.execute(f'UPDATE day_db{(function_call.data)[16:]} SET client{str(k+1)} = ? WHERE time = ?', (function_call.message.chat.id, ((function_call.data)[11:16] + ":00")))
                    connection.commit()
                    cursor.execute(f'SELECT client1, client2, client3, client4, client5, client6, client7, client8, client9, client10, client11, client12 FROM day_db{(function_call.data)[16:]} WHERE time = ?', (((function_call.data)[11:16] + ":00"),))
                    clients = cursor.fetchall()
                    fl = False
                    k += 1
                    break
                else:
                    k += 1
            if fl == True and i == 0:
                bot.edit_message_text(text="Ошибка!\U0000274C В компании больше нет мест. Пожалуйста, попробуйте найти новую компанию\U0001F50E", chat_id=function_call.message.chat.id, message_id=function_call.message.id)
                flag = False   
        if flag == True:
            cursor.execute('SELECT day FROM Shedule WHERE num = ?', ((function_call.data)[16:],))
            dating = cursor.fetchall()
            connection.close()
            connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
            cursor = connection.cursor()
            cursor.execute('SELECT nick, phone, name FROM Users WHERE id = ?', (function_call.message.chat.id,))
            user_info = cursor.fetchall()
            bot.send_message(chat_id=1333967466, text=f"Новая запись!\nДата: {dating[0][0]}\nВремя: {(function_call.data)[11:16]}\nНик в ТГ: {user_info[0][0]}\nТелефон: {user_info[0][1]}\nИмя: {user_info[0][2]}")
            bot.edit_message_text(text=f"Ждём вас {dating[0][0]} в {(function_call.data)[11:16]}! Приходите!\U00002705", chat_id=function_call.message.chat.id, message_id=function_call.message.id)
            markup = types.InlineKeyboardMarkup()
            button_book = types.InlineKeyboardButton(text="Забронировать стол\U0001F195\U00002705", callback_data="book")
            button_company = types.InlineKeyboardButton(text="Найти компанию🔎", callback_data="find_company")
            markup.add(button_book)
            markup.add(button_company)
            mess = "Для бронирования столика воспользуйтесь кнопками\U0001F4CD\n\nДля связи с оператором введите\n/help\U00002B07"
            bot.send_message(function_call.message.chat.id, mess, reply_markup=markup) 
            connection.close()
    if function_call.data == "view":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[:7]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_back = types.InlineKeyboardButton(text="Назад\U00002B05\U00002B05", callback_data="main")
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw1")
        markup.add(button_back, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw1":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[7:14]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw2")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="view")
        markup.add(button_prev, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw2":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[14:21]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw3")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="vw1")
        markup.add(button_prev, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw3":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[21:28]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw4")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="vw2")
        markup.add(button_prev, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw4":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[28:35]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw5")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="vw3")
        markup.add(button_prev, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw5":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[35:42]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw6")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="vw4")
        markup.add(button_prev, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw6":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[42:49]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="vw7")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="vw5")
        markup.add(button_prev, button_next)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "vw7":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[49:]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"vw_table{day[0]}")
            markup.add(button_day)
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="vw6")
        markup.add(button_prev)
        mess = "Для какого дня показать таблицу?"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "vw_table" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM day_db{(function_call.data)[8:]}")
        query = cursor.fetchall()
        data = []
        connection.close()
        for i in range(len(query)):
            datapr = []
            datapr.append(query[i][0])
            datapr.append(query[i][1])
            datapr.append(query[i][2])
            if query[i][3] == 0:
                connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
                cursor = connection.cursor()
                for j in range(4, 16):
                    if query[i][j] != None:
                        cursor.execute('SELECT phone, name FROM Users WHERE id = ?', (query[i][j],))
                        user_info = cursor.fetchall()
                        datapr.append(user_info[0][0])
                        datapr.append(user_info[0][1])
                connection.close()
            else:
                connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
                cursor = connection.cursor()
                for j in range(4, 16):
                    if query[i][j] != None:
                        cursor.execute('SELECT phone, name FROM UsersAdmin WHERE phone = ?', (query[i][j],))
                        user_info = cursor.fetchall()
                        datapr.append(str(user_info[0][0]))
                        datapr.append(user_info[0][1])
                connection.close()
            data.append(datapr)
        wb = Workbook()
        ws = wb.active
        ws.append(["Номер", "Время", "Оплата", "Телефон 1", "Имя 1", "Телефон 2", "Имя 2", "Телефон 3", "Имя 3", "Телефон 4", "Имя 4", "Телефон 5", "Имя 5", "Телефон 6", "Имя 6", "Телефон 7", "Имя 7", "Телефон 8", "Имя 8", "Телефон 9", "Имя 9", "Телефон 10", "Имя 10", "Телефон 11", "Имя 11", "Телефон 12", "Имя 12"])
        for i in range(len(data)):
            ws.append(data[i])
        for i in range(1, 16):
            ws[f'A{i}'].font = Font(bold=True)
        for i in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            ws[f'{i}1'].font = Font(bold=True)
        ws[f'AA1'].font = Font(bold=True)
        wb.save(f"C:/WORK/Bots/Mafia_reg_bot/day_db{(function_call.data)[8:]}.xlsx")
        file = open(f"C:/WORK/Bots/Mafia_reg_bot/day_db{(function_call.data)[8:]}.xlsx", "rb")
        bot.send_document(chat_id=function_call.message.chat.id, document=file)
        file.close()
    elif function_call.data == "view_all":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
        query = f"SELECT * FROM Users"
        df = pd.read_sql(query, connection)
        df.to_excel(f"C:/WORK/Bots/Mafia_reg_bot/Users.xlsx")
        file = open(f"C:/WORK/Bots/Mafia_reg_bot/Users.xlsx", "rb")
        bot.send_document(chat_id=function_call.message.chat.id, document=file)
        file.close() 
    elif function_call.data == "sending":
        msg = bot.send_message(function_call.message.chat.id, "Отправьте сообщение, которое нужно отправить всем пользователям бота!")
        bot.register_next_step_handler(msg, send_toall)
    elif function_call.data == "add_admin":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[:7]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_back = types.InlineKeyboardButton(text="Назад\U00002B05\U00002B05", callback_data="admin_main")
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_1")
        markup.add(button_back, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_1":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[7:14]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_2")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="add_admin")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_2":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[14:21]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_3")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_1")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_3":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[21:28]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_4")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_2")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_4":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[28:35]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_5")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_3")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_5":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[35:42]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_6")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_4")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_6":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[42:49]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_7")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_5")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "adm_7":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[49:]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_admn{day[0]}")
            markup.add(button_day)
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_6")
        markup.add(button_prev)
        mess = "На какое число нужно добавить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "day_admn" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT num, time, client1 from day_db{str((function_call.data)[8:])}')
        data = cursor.fetchall()
        cursor.execute('SELECT day FROM Shedule WHERE num = ?', (str(function_call.data)[8:],))
        day = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect(f'C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS Writes_admin (
        id INTEGER PRIMARY KEY,
        date TEXT,
        time TEXT,
        hours INTEGER,
        name TEXT,
        phone INTEGER
        )
        ''')
        try:
            cursor.execute('SELECT * FROM Writes_admin WHERE id = ?', (int(function_call.message.chat.id),))
        except:
            pass
        ifwrite = cursor.fetchall()
        if ifwrite != []:
            cursor.execute('DELETE FROM Writes_admin WHERE id = ?', (function_call.message.chat.id,))
        connection.commit()
        cursor.execute('INSERT INTO Writes_admin (id, date) VALUES (?, ?)', (function_call.message.chat.id, day[0][0]))
        connection.commit()
        markup = types.InlineKeyboardMarkup()
        spis_buttons = []
        for dat in range(len(data)):
            if data[dat][2] == None:
                count = 0
                if dat + 5 < len(data):
                    for i in range(1, 5):
                        if data[dat + i][2] == None:
                            count += 1
                        else:
                            break
                else:
                    fl = True
                    for i in range(1, (len(data)-dat)):
                        if data[dat + i][2] == None:
                            count += 1
                        else:
                            fl = False
                            break
                    if fl == True:
                        count = 4
                if count != 0:
                    spis_buttons.append(types.InlineKeyboardButton(text=f"{data[dat][1][:5]}, лимит: {str(count)} час(а)", callback_data=f"tmad{data[dat][1][:5]}{str(count)}"))
        for i in range(0, len(spis_buttons)-1, 2):
            markup.add(spis_buttons[i], spis_buttons[i+1])
        connection.close()    
        mess = "На какое время добавляем бронь?\U0001F551"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
    elif "tmad" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Writes_admin SET time = ? WHERE id = ?', ((function_call.data)[4:9], function_call.message.chat.id))
        connection.commit()
        markup = types.InlineKeyboardMarkup()
        for i in range(1, int((function_call.data)[9:]) + 1):
            button = types.InlineKeyboardButton(text=f"{i} час(а)", callback_data=f"tmd_lst{i}")
            markup.add(button)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text="Сколько часов бронируем?\U0001F550", reply_markup=markup, message_id=function_call.message.id)
    elif "tmd_lst" in function_call.data:
        connection = sqlite3.connect(f'C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Writes_admin SET hours = ? WHERE id = ?', ((function_call.data)[7:], function_call.message.chat.id))
        connection.commit()
        mess = "Введите номер телефона клиента\U0001F4DE\n\nПример: +79259998877"
        msg = bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, message_id=function_call.message.id)
        bot.register_next_step_handler(msg, add_name)
    elif function_call.data == "admin_payment":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * FROM Writes_admin WHERE id = ?', (function_call.message.chat.id,))
        reg = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT num from Shedule WHERE day = ?', (reg[0][1],))
        num = cursor.fetchall()
        num = num[0][0]
        cursor.execute(f'SELECT num from day_db{num} WHERE time = ?', ((reg[0][2] + ":00"),))
        num_time = cursor.fetchall()
        num_time = num_time[0][0]
        cursor.execute(f'UPDATE day_db{num} SET pay = ? WHERE num = ?', (True, (num_time)))
        cursor.execute(f'UPDATE day_db{num} SET administrator = ? WHERE num = ?', (True, (num_time)))
        for i in range(reg[0][3]):
            cursor.execute(f'UPDATE day_db{num} SET pay = ? WHERE num = ?', (True, (num_time + i)))
            cursor.execute(f'UPDATE day_db{num} SET administrator = ? WHERE num = ?', (True, (num_time + i)))
            cursor.execute(f'UPDATE day_db{num} SET client1 = ? WHERE num = ?', (reg[0][5], (num_time + i)))
        connection.commit()
        num_time += (i+1)
        for i in range(num_time, 15):
            cursor.execute(f'SELECT time, client1 from day_db{num} WHERE num = ?', (i,))
            current_reg = cursor.fetchall()
            if current_reg[0][1] == None:
                c_time = current_reg[0][0]
                hr = int(c_time[:2])
                mnte = int(c_time[3:5])
                if mnte != 40:
                    mnte += 20
                else:
                    mnte = 0
                    hr += 1
                if mnte == 0:
                    mnte = "00"
                times = str(hr) + ":" + str(mnte) + ":" + "00"
                cursor.execute(f'UPDATE day_db{num} SET time = ? WHERE num = ?', (times, i))
                connection.commit()
            else:
               break
        cursor.execute('SELECT day FROM Shedule WHERE num = ?', (num,))
        dating = cursor.fetchall()
        mess = f"Ждём вас {reg[0][1]} в {reg[0][2]} на {reg[0][3]} час(а)\U0001F917"
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
        cursor = connection.cursor()
        cursor.execute('SELECT phone, name FROM UsersAdmin WHERE phone = ?', (reg[0][5],))
        user_info = cursor.fetchall()
        bot.send_message(chat_id=1333967466, text=f"Новая запись!\nДата: {dating[0][0]}\nВремя: {reg[0][2]}\nТелефон: {user_info[0][0]}\nИмя: {user_info[0][1]}")
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, message_id=function_call.message.id)
        markup = types.InlineKeyboardMarkup()
        button1 = types.InlineKeyboardButton(text="Рассылка", callback_data="sending")
        button = types.InlineKeyboardButton(text="Посмотреть записи", callback_data="view")
        button2 = types.InlineKeyboardButton(text="Все пользователи", callback_data="view_all")
        button3 = types.InlineKeyboardButton(text="Добавить бронь", callback_data="add_admin")
        button4 = types.InlineKeyboardButton(text="Удалить бронь", callback_data="del_admin")
        markup.add(button1)
        markup.add(button, button2)
        markup.add(button3, button4)
        bot.edit_message_text(chat_id=function_call.message.chat.id, text="Меню администратора\U00002B07", reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "del_admin":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[:7]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_back = types.InlineKeyboardButton(text="Назад\U00002B05\U00002B05", callback_data="admin_main")
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="mda_1")
        markup.add(button_back, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_1":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[7:14]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="mda_2")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="add_admin")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_2":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[14:21]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="mda_3")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="mda_1")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_3":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[21:28]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="mda_4")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="mda_2")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_4":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[28:35]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="adm_5")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="adm_3")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_5":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[35:42]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="mda_6")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="mda_4")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_6":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[42:49]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_next = types.InlineKeyboardButton(text="\U000027A1", callback_data="mda_7")
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="mda_5")
        markup.add(button_prev, button_next)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif function_call.data == "mda_7":
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * from Shedule')
        days = cursor.fetchall()
        markup = types.InlineKeyboardMarkup()
        for day in days[49:]:
            button_day = types.InlineKeyboardButton(text=f"{day[2]}, {day[1][:5]}", callback_data=f"day_mnda{day[0]}")
            markup.add(button_day)
        button_prev = types.InlineKeyboardButton(text="\U00002B05", callback_data="mda_6")
        markup.add(button_prev)
        mess = "На какое число нужно удалить бронь?\U0001F4C6"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
        connection.close()
    elif "day_mnda" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f'SELECT num, time, client1 from day_db{str((function_call.data)[8:])}')
        data = cursor.fetchall()
        cursor.execute('SELECT day FROM Shedule WHERE num = ?', (str(function_call.data)[8:],))
        day = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect(f'C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS Writes_admin (
        id INTEGER PRIMARY KEY,
        date TEXT,
        time TEXT,
        hours INTEGER,
        name TEXT,
        phone INTEGER
        )
        ''')
        try:
            cursor.execute('SELECT * FROM Writes_admin WHERE id = ?', (int(function_call.message.chat.id),))
        except:
            pass
        ifwrite = cursor.fetchall()
        if ifwrite != []:
            cursor.execute('DELETE FROM Writes_admin WHERE id = ?', (function_call.message.chat.id,))
        connection.commit()
        cursor.execute('INSERT INTO Writes_admin (id, date) VALUES (?, ?)', (function_call.message.chat.id, day[0][0]))
        connection.commit()
        markup = types.InlineKeyboardMarkup()
        spis_buttons = []
        for dat in range(len(data)):
            spis_buttons.append(types.InlineKeyboardButton(text=f"{data[dat][1][:5]}", callback_data=f"mtad{data[dat][1][:5]}{str((function_call.data)[8:])}"))
        for i in range(0, len(spis_buttons)-1, 2):
            markup.add(spis_buttons[i], spis_buttons[i+1])
        connection.close()    
        mess = "На какое время удаляем бронь?\U0001F551"
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=mess, reply_markup=markup, message_id=function_call.message.id)
    elif "mtad" in function_call.data:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('SELECT * FROM Writes_admin WHERE id = ?', (function_call.message.chat.id,))
        data = cursor.fetchall()
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
        cursor = connection.cursor()
        cursor.execute(f'UPDATE day_db{str((function_call.data)[9:])} SET pay = NULL, administrator = NULL, client1 = NULL, client2 = NULL, client3 = NULL, client4 = NULL, client5 = NULL, client6 = NULL, client7 = NULL, client8 = NULL, client9 = NULL, client10 = NULL, client11 = NULL, client12 = NULL WHERE time = ?', ((function_call.data)[4:9] + ":00",))
        connection.commit()
        connection.close()
        bot.edit_message_text(chat_id=function_call.message.chat.id, text=f"Запись на {data[0][1]} в {str((function_call.data)[4:9])} удалена!\U00002705", message_id=function_call.message.id)       
        markup = types.InlineKeyboardMarkup()
        button1 = types.InlineKeyboardButton(text="Рассылка", callback_data="sending")
        button = types.InlineKeyboardButton(text="Посмотреть записи", callback_data="view")
        button2 = types.InlineKeyboardButton(text="Все пользователи", callback_data="view_all")
        button3 = types.InlineKeyboardButton(text="Добавить бронь", callback_data="add_admin")
        button4 = types.InlineKeyboardButton(text="Удалить бронь", callback_data="del_admin")
        markup.add(button1)
        markup.add(button, button2)
        markup.add(button3, button4)
        bot.send_message(chat_id=function_call.message.chat.id, text="Меню администратора\U00002B07", reply_markup=markup)  
    elif function_call.data == "see_month":
        bot.send_message(function_call.message.chat.id, "\U0001F551Подождите немного. Таблица создаётся")
        wb = Workbook()
        for dy in range(1,32):
            connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
            cursor = connection.cursor()
            cursor.execute('SELECT day FROM Shedule WHERE num = ?', (dy,))
            day = cursor.fetchall()
            cursor.execute(f"SELECT * FROM day_db{dy}")
            query = cursor.fetchall()
            data = []
            connection.close()
            for i in range(len(query)):
                datapr = []
                datapr.append(query[i][0])
                datapr.append(query[i][1])
                datapr.append(query[i][2])
                if query[i][3] == 0:
                    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
                    cursor = connection.cursor()
                    for j in range(4, 16):
                        if query[i][j] != None:
                            cursor.execute('SELECT phone, name FROM Users WHERE id = ?', (query[i][j],))
                            user_info = cursor.fetchall()
                            datapr.append(user_info[0][0])
                            datapr.append(user_info[0][1])
                    connection.close()
                else:
                    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
                    cursor = connection.cursor()
                    for j in range(4, 16):
                        if query[i][j] != None:
                            cursor.execute('SELECT phone, name FROM UsersAdmin WHERE phone = ?', (query[i][j],))
                            user_info = cursor.fetchall()
                            print(user_info)
                            datapr.append(str(user_info[0][0]))
                            datapr.append(user_info[0][1])
                    connection.close()
                data.append(datapr)
            ws = wb.create_sheet(f"{day[0][0]}")
            ws.append(["Номер", "Время", "Оплата", "Телефон 1", "Имя 1", "Телефон 2", "Имя 2", "Телефон 3", "Имя 3", "Телефон 4", "Имя 4", "Телефон 5", "Имя 5", "Телефон 6", "Имя 6", "Телефон 7", "Имя 7", "Телефон 8", "Имя 8", "Телефон 9", "Имя 9", "Телефон 10", "Имя 10", "Телефон 11", "Имя 11", "Телефон 12", "Имя 12"])
            for i in range(len(data)):
                ws.append(data[i])
            for i in range(1, 16):
                ws[f'A{i}'].font = Font(bold=True)
            for i in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                ws[f'{i}1'].font = Font(bold=True)
            ws[f'AA1'].font = Font(bold=True)
        del wb["Sheet"]
        wb.save(f"C:/WORK/Bots/Mafia_reg_bot/month_days.xlsx")
        file = open(f"C:/WORK/Bots/Mafia_reg_bot/month_days.xlsx", "rb")
        bot.send_document(chat_id=function_call.message.chat.id, document=file)
        file.close()


def send_toall(message):
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM Users")
    users = cursor.fetchall()
    for user in users:
        try:
            bot.copy_message(int(user[0]), message.chat.id, message.id)
        except:
            cursor.execute('DELETE FROM Users WHERE id = ?', (user[0],))
    connection.commit()
    connection.close()

def add_name(message):
    if (message.text)[0] != "+":
        msg = bot.send_message(message.chat.id, "\U00002757\U00002757Номер телефона введён неверно. Введите номер телефона ещё раз")
        bot.register_next_step_handler(msg, add_name)
    else:
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/temporary.db')
        cursor = connection.cursor()
        cursor.execute('UPDATE Writes_admin SET phone = ? WHERE id = ?', (message.text, message.chat.id))
        connection.commit()
        connection.close()
        connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
        cursor = connection.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS UsersAdmin (
        phone INTEGER,
        name TEXT
        )
        ''')
        cursor.execute('SELECT * FROM UsersAdmin WHERE phone = ?', (int(message.text),))
        info = cursor.fetchall()
        if info == []:
            mess = bot.send_message(message.chat.id, "Введите имя клиента")
            connection.close()
            bot.register_next_step_handler(mess, add_final_name, message.text, message)
        else:
            connection = sqlite3.connect(f'C:/WORK/Bots/Mafia_reg_bot/temporary.db')
            cursor = connection.cursor()
            cursor.execute('SELECT * FROM Writes_admin WHERE id = ?', (message.chat.id,))
            data = cursor.fetchall()
            mess = f"Проверьте правильность записи:\U0001F50E\n\nДата: {data[0][1]}\nВремя: {data[0][2][:5]}\nКоличество часов: {data[0][3]}\nТелефон: {info[0][0]}\nИмя: {info[0][1]}"
            connection.close()
            markup = types.InlineKeyboardMarkup()
            button_yes = types.InlineKeyboardButton(text="Всё верно\U00002705", callback_data="admin_payment")
            button_back = types.InlineKeyboardButton(text="Ошибка\U0000274C", callback_data="admin_main")
            markup.add(button_yes)
            markup.add(button_back)
            bot.send_message(chat_id=message.chat.id, text=mess, reply_markup=markup)

def add_final_name(message, phone, msg):
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Users_info.db')
    cursor = connection.cursor()
    cursor.execute("INSERT INTO UsersAdmin (phone, name) VALUES (?, ?)", (phone, message.text))
    connection.commit()
    connection.close()
    add_name(msg)



def start_polling():
    bot.infinity_polling(none_stop=True)

polling_thread = threading.Thread(target=start_polling)
polling_thread.start()

def command():
    print("command")
    spis = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    cur_date = date.today()
    cur_date = cur_date + timedelta(days=56)
    week = cur_date
    cur_date = str(cur_date)
    cur_date = cur_date[8:] + "." + cur_date[5:7] + "." + cur_date[2:4]
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
    cursor = connection.cursor()
    cursor.execute('INSERT INTO Shedule (day, week) VALUES (?, ?)', (cur_date, spis[week.weekday()]))
    connection.commit()
    cursor.execute('SELECT num FROM Shedule WHERE day = ?', (cur_date,))
    num = cursor.fetchall()
    cursor.execute(f'''
    CREATE TABLE IF NOT EXISTS day_db{num[0][0]} (
    num INTEGER PRIMARY KEY AUTOINCREMENT,
    time TEXT,
    pay BOOLEAN,
    administrator BOOLEAN,
    client1 TEXT,
    client2 TEXT,
    client3 TEXT,
    client4 TEXT,
    client5 TEXT,
    client6 TEXT,
    client7 TEXT,
    client8 TEXT,
    client9 TEXT,
    client10 TEXT,
    client11 TEXT,
    client12 TEXT
    )
    ''')
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(10, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(11, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(12, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(13, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(14, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(15, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(16, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(17, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(18, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(19, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(20, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(21, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(22, 0, 0)),))
    cursor.execute(f'INSERT INTO day_db{num[0][0]} (time) VALUES (?)', (str(time(23, 0, 0)),))
    connection.commit()
    connection.close()
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Find_company.db')
    cursor = connection.cursor()
    cursor.execute(f'''
    CREATE TABLE IF NOT EXISTS find_db{num[0][0]} (
    num INTEGER PRIMARY KEY AUTOINCREMENT,
    time TEXT,
    client1 TEXT,
    client2 TEXT,
    client3 TEXT,
    client4 TEXT,
    client5 TEXT,
    client6 TEXT,
    client7 TEXT,
    client8 TEXT,
    client9 TEXT,
    client10 TEXT,
    client11 TEXT,
    client12 TEXT
    )
    ''')
    cursor.execute(f'INSERT INTO find_db{num[0][0]} (time) VALUES (?)', ("Утро",))
    cursor.execute(f'INSERT INTO find_db{num[0][0]} (time) VALUES (?)', ("День",))
    cursor.execute(f'INSERT INTO find_db{num[0][0]} (time) VALUES (?)', ("Вечер",))
    connection.commit()
    connection.close()
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Days_work.db')
    cursor = connection.cursor()
    cursor.execute('DELETE FROM Shedule WHERE num = ?', (1,))
    cursor.execute('DROP TABLE IF EXISTS day_db1')
    connection.commit()
    for i in range(56):
        cursor.execute(f'ALTER TABLE day_db{i+2} RENAME TO day_db{i+1}')
        cursor.execute(f'UPDATE Shedule SET num = ? WHERE num = ?', (i+1, i+2))
    connection.commit()
    connection.close()
    connection = sqlite3.connect('C:/WORK/Bots/Mafia_reg_bot/Find_company.db')
    cursor = connection.cursor()
    cursor.execute('DROP TABLE IF EXISTS find_db1')
    connection.commit()
    for i in range(55):
        cursor.execute(f'ALTER TABLE find_db{i+2} RENAME TO find_db{i+1}')
    connection.commit()
    connection.close()
    
schedule.every().day.at("00:01").do(command)

while True:
    schedule.run_pending()
    tm.sleep(1)